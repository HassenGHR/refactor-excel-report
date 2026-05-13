#!/usr/bin/env python3
"""
to_router_excel.py — Build an Excel file optimized for *maximum data extraction*
by the existing excel_import_service.py (TP parser path), without needing any
change to the FastAPI route.

Pipeline:
    source.xlsx  ──► parse_ddr() (reads everything we can find)
                 ──► build_router_excel() (writes a layout the TP parser scans)
                 ──► output.zip (containing Excel)  ──► unzip ──► Excel ──► POST /v1/excel_import/upload  ──► full dict

The output Excel does NOT preserve any visual template. Its job is purely to
expose every extracted field at the cell positions and label keywords that
the original `_parse_tp`, `_extract_operations`, `_extract_mud_properties`,
`_extract_mud_volume`, `_extract_mud_chemicals`, and `_extract_personnel`
functions will find.

CLI:
    python to_router_excel.py SOURCE.xlsx [-o OUTPUT.zip]
"""
from __future__ import annotations
import argparse
import io
import re
import sys
import zipfile
from datetime import date as date_type, datetime
from pathlib import Path
from openpyxl import Workbook

# Re-use the source extractor from ddr_extract.py (must be in same folder)
from helpers.parse_source import parse_source as parse_ddr


# Non-chemical keywords from excel_import_service._extract_mud_chemicals.
# We use this set to PREFIX personnel labels with a token that matches one of
# these keywords, so personnel rows in the W-Z area (rows 25-50) don't get
# misclassified as mud chemicals (which scans the same area, rows 15-40).
PERSONNEL_PREFIX = "PERSONNEL: "    # "PERSONNEL" is in the non-chemical list


# The existing TP parser scans cols 23-26 rows 25-50 for SUPERVISOR-flavoured
# keywords. If a personnel role name contains one of these, the scan will
# false-match and pick up the count column as the supervisor's name. We rewrite
# trigger words to a safe equivalent before writing the personnel cell.
_TRIGGER_SUBS = [
    (re.compile(r"\bMA[IÎ]TRE\b",          re.IGNORECASE), "M.O."),
    (re.compile(r"\bREPR[ÉE]SENTANT\b",    re.IGNORECASE), "Repr."),
    (re.compile(r"\bSUPERVIS(?:EUR|OR)\b", re.IGNORECASE), "Supv."),
    (re.compile(r"SH/DP",                   re.IGNORECASE), "SHDP"),
]
def _safe_personnel_label(name: str) -> str:
    for pat, repl in _TRIGGER_SUBS:
        name = pat.sub(repl, name)
    return name


def build_router_excel(data: dict, output_path: Path) -> None:
    """Lay out every extractable field where the TP parser will find it."""
    wb = Workbook()
    ws = wb.active

    h = data.get("header", {})

    rd = h.get("date")
    ws.title = (rd.strftime("%d-%m-%Y")
                if isinstance(rd, (datetime, date_type)) else "report")

    # =========================================================================
    # Title — triggers _detect_format() → 'tp'
    # =========================================================================
    ws["A1"] = "Rapport journalier de Work-over"

    # =========================================================================
    # Report metadata in the upper-right (rows 1-7, cols 20-28)
    # =========================================================================
    if rd:
        ws.cell(4, 22).value = rd                    # V4 — picked up by date scan
    if h.get("day_number"):
        ws.cell(4, 25).value = "N°"                  # Y4 label
        ws.cell(4, 26).value = h["day_number"]       # Z4 value

    # =========================================================================
    # Well / Field / Rig (row 6 — label-based scan)
    # =========================================================================
    ws["A6"] = "Puits"
    ws["B6"] = h.get("well_name", "")
    ws["D6"] = "Champ"
    ws["E6"] = h.get("field_name", "")
    ws["G6"] = "Appareil"
    ws["I6"] = h.get("rig_name", "")

    # =========================================================================
    # Casing / BHA — Dernier Tubage, Top Liner labels
    # Parser reads: csg_info = label+1, csg_depth = label+2 or +3
    # =========================================================================
    ws["A7"] = "Dernier Tubage"
    if h.get("last_csg_size"):
        ws["B7"] = h["last_csg_size"]
    if h.get("last_csg_depth"):
        ws["C7"] = h["last_csg_depth"]

    if h.get("last_csg_top"):
        # The TP parser reads `_cell(ws, pos[0], pos[1])` at the label cell
        # itself, so the label *must* contain the value.
        ws["A8"] = f'Top Liner {h["last_csg_top"]}'

    # Extra context (top shoe, BHA details) — written with stable labels so
    # the parser-side patch can pick them up.
    if h.get("top_shoe"):
        ws["A9"] = "Top Shoe"          # parser scans for "TOP SHOE" → last_csg_shoe_real / lastCSNlnrSHOE
        ws["B9"] = h["top_shoe"]
    if h.get("bha_details"):
        ws["A10"] = "BHA"
        ws["B10"] = h["bha_details"]

    # =========================================================================
    # Operations table — headers at row 17, data starting row 18
    # The dynamic detector finds: DE/A header row, then Code & Heure labels,
    # then 'OPERATIONS' for the description start column.
    # =========================================================================
    ws.cell(17, 1).value  = "DE"
    ws.cell(17, 2).value  = "A"
    ws.cell(17, 3).value  = "OPERATIONS"     # desc start
    ws.cell(17, 13).value = "Code"           # tarif col
    ws.cell(17, 14).value = "Heure"          # hours col

    activities = data.get("activities", [])
    for i, op in enumerate(activities):
        row = 18 + i
        ws.cell(row, 1).value  = op.get("start_time")
        ws.cell(row, 2).value  = op.get("end_time")
        ws.cell(row, 3).value  = op.get("description", "")
        ws.cell(row, 13).value = op.get("bill", "")
        ws.cell(row, 14).value = op.get("hours", 0)

    # Leave at least 3 empty rows after the last operation so the parser's
    # "stop after 3 empty rows" rule fires cleanly before any subsequent text.
    text_row = 18 + len(activities) + 3

    # =========================================================================
    # Text sections — scanned in rows 20-55, cols 1/3/4
    # =========================================================================
    sec = data.get("text_sections", {}) or {}
    safety = data.get("safety", {}) or {}

    if sec.get("after_midnight"):
        ws.cell(text_row, 1).value = "APRÈS MINUIT"
        ws.cell(text_row + 1, 3).value = sec["after_midnight"]
        # Parser scans 5 rows after the label for cols 3-12; keep next section
        # at least 6 rows away to avoid bleed-through.
        text_row += 6

    if sec.get("current_operation") or sec.get("day_summary"):
        ws.cell(text_row, 1).value = "Situation"           # 'SITUATION' keyword
        ws.cell(text_row, 4).value = sec.get("current_operation") or sec.get("day_summary")
        text_row += 2

    if sec.get("plan_operations"):
        ws.cell(text_row, 1).value = "Programme prévu"     # 'PROGRAMME' keyword
        ws.cell(text_row, 4).value = sec["plan_operations"]
        text_row += 2

    # =========================================================================
    # BOP test — scanned in rows 25-50, cols 10-25 for "BOP" then dates around
    # =========================================================================
    if h.get("bop_test"):
        bop_row = 45     # safely past operations + text sections
        ws.cell(bop_row, 10).value = "BOP Test Date"
        ws.cell(bop_row, 11).value = h["bop_test"]

    # =========================================================================
    # Mud properties — rows 5-30, cols 23-27 (W-AA)
    # =========================================================================
    mud = data.get("mud_checks", {}) or {}
    mud_row = 5

    # Each tuple: (label_text, dict_key). Labels chosen to match parser patterns.
    mud_layout = [
        ("Densité",          "density"),    # → density
        ("V.Marsh",          "fun_vis"),    # → fun_vis
        ("Filtrat",          "apl_fl"),     # → apl_fl
        ("Stability ES",     "es"),         # → es (STABIL or ES)
        ("Solid %",          "solid"),      # → solid (must contain SOLID and %)
        ("Huile",            "oil"),        # → oil
        ("H2O",              "h2o"),        # → h2o
        ("PV",               "pv"),         # → pv
        ("YP",               "yp"),         # → yp
        ("Gel 10sec",        "gel10sec"),   # → gel10sec
        ("Gel 10m",          "gel10m"),     # → gel10m
        ("PH",               "ph"),         # → ph
        ("POM",              "pom"),        # → pom
        ("ECD",              "ecd"),        # → ecd
        ("Sand",             "sand"),       # → sand
        ("LGS",              "lgs"),        # → lgs
        ("HGS",              "hgs"),        # → hgs
        ("NaCl",             "NaCl"),       # → NaCl
        ("HPHT",             "hpht_fl"),    # → hpht_fl
        ("E LIME",           "e_line"),     # → e_line
        ("MBT",              "mbt"),        # → mbt
    ]
    for label, key in mud_layout:
        v = mud.get(key)
        if v is None or v == "":
            continue
        ws.cell(mud_row, 23).value = label    # W col
        ws.cell(mud_row, 24).value = v        # X col (label+1)
        mud_row += 1

    # Oil/Water ratio — separate, text-valued
    if mud.get("oil_water_ratio"):
        ws.cell(mud_row, 23).value = "H/E"
        ws.cell(mud_row, 24).value = mud["oil_water_ratio"]
        mud_row += 1

    # =========================================================================
    # Mud volume — must be within rows 5-15 to match the parser's scan range
    # =========================================================================
    vol = data.get("mud_volume", {}) or {}
    # Use rows 13-15 for volumes (inside the (5, 15) scan range)
    vol_row = 13
    if vol.get("string_volume") is not None:
        ws.cell(vol_row, 23).value = "Volume Puits (m³)"
        ws.cell(vol_row, 24).value = vol["string_volume"]
        vol_row += 1
    if vol.get("pits_volume") is not None:
        ws.cell(vol_row, 23).value = "Volume Surface"
        ws.cell(vol_row, 24).value = vol["pits_volume"]
        vol_row += 1
    if vol.get("total_volume") is not None:
        ws.cell(vol_row, 23).value = "Total Volume"     # TOTAL only, no SURFACE/PUITS
        ws.cell(vol_row, 24).value = vol["total_volume"]
        vol_row += 1

    # =========================================================================
    # Mud chemicals — chemical scan zone is rows 15-40 col 23.
    # Rows 16-24: pure chemical zone (col 23 = item, col 24 = stock).
    # Rows 25+: shared with personnel zone, so we leave col 24 empty (so
    #          personnel scan skips this row — count_val is None) and put
    #          stock value in col 25, which the parser's `or` fallback reads.
    # =========================================================================
    PURE_CHEM_START, PURE_CHEM_END, CHEM_SCAN_END = 16, 24, 40
    chem_row = PURE_CHEM_START
    chemicals = data.get("mud_chemical_usage", []) or []
    for chem in chemicals:
        item = (chem.get("item") or "").strip()
        if not item:
            continue
        if chem_row > CHEM_SCAN_END:
            break
        stock = (chem.get("on_loc")
                 or chem.get("used")
                 or chem.get("received")
                 or "0")
        ws.cell(chem_row, 23).value = item
        if chem_row <= PURE_CHEM_END:
            ws.cell(chem_row, 24).value = stock
        else:
            # Shared zone: stock goes in col 25, col 24 stays empty so
            # the personnel scan skips this row.
            ws.cell(chem_row, 25).value = stock
        chem_row += 1

    # =========================================================================
    # Personnel — rows 25-44, cols 23/24/25
    # Use PERSONNEL: prefix so the same rows aren't also classified as chemicals
    # (the chemical scan filters labels containing "PERSONNEL").  Start after
    # any chemical rows we wrote into the shared zone, so they don't collide.
    # =========================================================================
    p_row = max(25, chem_row)
    for p in data.get("personnel_data", []):
        company = p.get("company", "").strip()
        if not company:
            continue
        if p_row > 44:                       # leave 45-50 for supervisor / BOP
            break
        safe_company = _safe_personnel_label(company)
        ws.cell(p_row, 23).value = f"{PERSONNEL_PREFIX}{safe_company}"
        ws.cell(p_row, 24).value = p.get("number", 0)
        if p.get("names"):
            ws.cell(p_row, 25).value = p["names"]
        p_row += 1

    # =========================================================================
    # Supervisor / Junior Supervisor — picked by the supervisor scan that
    # looks for MAITRE/REPRÉSENTANT/SUPERVISEUR/SUPERVISOR/SH/DP in W-AA rows 25-50
    # =========================================================================
    # Both names on a rig report are 12-hour-shift supervisors.  The DB has
    # both a `supervisor` and a `superintendent` column, so we use them as
    # slot 1 and slot 2 — "superintendent" is a column-name misnomer, the
    # second person is really a second supervisor.  We write at the END of
    # the W column (rows 46+) so this overrides any false matches from
    # personnel rows whose pre-mangling names contained trigger words.
    # No fallback between the two — each is written only when the extractor
    # actually set it.
    sup_row = max(p_row, 46)
    if h.get("supervisor"):
        ws.cell(sup_row, 23).value = "Représentant SH/DP"
        ws.cell(sup_row, 24).value = h["supervisor"]
        sup_row += 1
    if h.get("superintendent"):
        # Parser-side patch scans for "SUPERINTEND" → header.superintendent.
        ws.cell(sup_row, 23).value = "Superintendent"
        ws.cell(sup_row, 24).value = h["superintendent"]
        sup_row += 1

    # =========================================================================
    # Extra labels — these aren't read by the CURRENT TP parser, but if you
    # ever patch the parser to extract `superintendent`, `well_md`, `remarks`,
    # `safety_data` etc., the data will already be in the Excel at predictable
    # cells, so the parser change is a tiny scan-for-label addition.
    # =========================================================================

    # Total Depth (visible cell for human + future parser hook)
    # Different rigs use different dict keys for this — check all of them.
    depth_value = (h.get("tmd")
                   or h.get("present_depth")
                   or h.get("well_md")
                   or h.get("total_depth"))
    if depth_value:
        ws["A11"] = "Total Depth"
        ws["B11"] = depth_value

    # NOTE: We do NOT write a "Superintendent" label here.
    # TP Junior is a supervisor, not the Superintendent — the Superintendent
    # is a separate role (the operator's company-man / SH-DP representative)
    # and its name comes from a different cell in the source DDR (TBD).
    # When that source is identified, set header["superintendent"] in
    # ddr_extract.py and write it here with the label "Superintendent".

    # Remark — written near the bottom in a clearly labelled row
    rmk_row = 60
    if sec.get("remarks") or sec.get("remark"):
        ws.cell(rmk_row, 1).value = "Remark"
        ws.cell(rmk_row, 3).value = sec.get("remarks") or sec.get("remark")
        rmk_row += 2

    # Accident incidents / safety data
    if safety.get("accident_topics") or sec.get("safety_data"):
        ws.cell(rmk_row, 1).value = "Accident Incident Topics"
        ws.cell(rmk_row, 3).value = safety.get("accident_topics") or sec.get("safety_data")
        rmk_row += 2

    # =========================================================================
    # HSE counters — written with explicit labels the parser-side patch scans.
    # =========================================================================
    if safety.get("accident_free_days") is not None:
        ws.cell(rmk_row, 1).value = "Accident Free Days"      # parser → header.accident_free_days
        ws.cell(rmk_row, 2).value = safety["accident_free_days"]
        rmk_row += 1
    if safety.get("water_truck") is not None:
        ws.cell(rmk_row, 1).value = "Water Truck"             # parser → header.water_truck
        ws.cell(rmk_row, 2).value = safety["water_truck"]
        rmk_row += 1
    if safety.get("accident_today") is not None:
        ws.cell(rmk_row, 1).value = "Accident Today"          # parser → header.accident_today
        ws.cell(rmk_row, 2).value = safety["accident_today"]
        rmk_row += 1

    # =========================================================================
    # Other safety counters (HSE meetings, permits, stop cards, exercises)
    # — kept for human reference; outside parser scan windows.
    # =========================================================================
    safe_row = max(rmk_row, 70)
    for label, key in [
        ("HSE meetings",       "hse_meetings"),
        ("Permits to work",    "permits_to_work"),
        ("Stop cards",         "stop_cards"),
        ("Safety exercises",   "exercises"),
    ]:
        if key in safety:
            ws.cell(safe_row, 1).value = label
            ws.cell(safe_row, 2).value = safety[key]
            safe_row += 1

    # =========================================================================
    # Tarif totals (extra info — daily T1/T2/T3/T4 hour totals)
    # =========================================================================
    totals = data.get("tarif_totals", {}) or {}
    if totals:
        ws.cell(safe_row + 1, 1).value = "Tarif daily totals (hours)"
        col = 2
        for code, hrs in sorted(totals.items()):
            ws.cell(safe_row + 1, col).value = code
            ws.cell(safe_row + 1, col + 1).value = hrs
            col += 2

    # Save Excel to memory, then zip it
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    with zipfile.ZipFile(output_path, 'w') as zf:
        zf.writestr(output_path.with_suffix('.xlsx').name, bio.getvalue())


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="Convert a source Excel report to a router-friendly Excel "
                    "the existing excel_import_service.py can fully parse."
    )
    p.add_argument("source", type=Path, help="Path to the source .xlsx file")
    p.add_argument("-o", "--output", type=Path, default=None,
                   help="Output .zip path (default: <rig>_<date>_router.zip)")
    args = p.parse_args(argv)

    if not args.source.exists():
        sys.exit(f"ERROR: source file not found: {args.source}")

    # 1. Extract everything we can from the source
    data = parse_ddr(args.source)

    # 2. Build the output path
    if args.output is None:
        rig = (data["header"].get("rig_name") or "rig").replace("/", "_")
        rd = data["header"].get("date")
        date_str = (rd.strftime("%Y-%m-%d")
                    if isinstance(rd, (datetime, date_type)) else "out")
        args.output = Path.cwd() / f"{rig}_{date_str}_router.zip"

    # 3. Lay everything out for the TP parser
    build_router_excel(data, args.output)

    # 4. Summary
    print(f"Wrote {args.output}")
    print(f"  Activities          : {len(data['activities'])}")
    print(f"  Personnel rows      : {len(data['personnel_data'])}")
    print(f"  Mud properties      : {len(data['mud_checks'])}")
    print(f"  Mud volume fields   : {len(data['mud_volume'])}")
    print(f"  Text sections       : {len(data['text_sections'])}")
    print(f"  Safety counters     : {len(data.get('safety', {}))}")
    print(f"  Tarif daily totals  : {data.get('tarif_totals', {})}")
    return 0


if __name__ == "__main__":
    sys.exit(main())