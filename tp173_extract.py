#!/usr/bin/env python3
"""
tp173_extract.py — extract a TP-173 (ENTP rig 173, Direction Régionale
ADRAR, ODZ wells) Daily Workover Report into the standard dict shape.

Source layout
-------------
Same template family as TP-179 ("RAPPORT JOURNALIER DU WORK OVER", 28 cols,
~40 rows, Adrar region) but with several positional differences:

  Field                 | TP-179            | TP-173
  ----------------------+-------------------+--------------------
  Title cell            | H2                | I2
  Well label            | "Well : HTJW 3"   | "ODZ-13" (bare, no label)
  Rig label             | "RIG : TP-179"    | "APPAREIL: TP-173"
  BOP test (real date)  | K7                | K6 inline (often placeholder)
  Last CSG SHOE         | -                 | K7 "DERNIER TUBAGE: Csg 7\"shoe @ 850m..."
  Activities first row  | 13 (cols A/C/D)   | 13 (cols A/B/C/D — extra start col)
  Personnel rows        | 28-33             | 27-31 (one up)
  Supervisor name       | S36               | S35 inline "Représentant maître d'œuvre :NAME"
  Water truck row       | r36 R=count       | (none — not on this template)

Distinguishing marker: title row contains "TP-173" specifically; or sheet has
"APPAREIL:" label (TP-179 uses "RIG :" instead).
"""
from __future__ import annotations
import re
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers (same shape as the other extractors)
# ---------------------------------------------------------------------------
def _clean(v) -> str:
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _float(v, default=0.0) -> float:
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    s = _clean(v).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "None"): return default
    s = re.sub(r"[a-zA-Zé°%/³]+$", "", s).strip()
    try: return float(s)
    except ValueError: return default


def _int(v, default=0) -> int:
    return int(_float(v, float(default)))


def _date_parse(v):
    """Parse a date; return None for placeholders like '--/--/2026' or '/'."""
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date_type): return v
    s = _clean(v)
    if not s or s in ("-", "/", "//"): return None
    if "--/--/" in s or "/--/" in s: return None       # placeholder
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None


def _time_parse(v):
    if v is None: return None
    if isinstance(v, time): return v
    if isinstance(v, datetime):
        # Excel epoch (1900-01-01) means end-of-day / 24:00
        return v.time()
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        if total == 86400: return time(0, 0)
        return time((total // 3600) % 24, (total % 3600) // 60)
    s = _clean(v)
    if not s or s in ("-", "None"): return None
    if s in ("24:00", "24:00:00"): return time(0, 0)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try: return datetime.strptime(s, fmt).time()
        except ValueError: continue
    return None


def _duration_hours(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, timedelta): return v.total_seconds() / 3600.0
    if isinstance(v, time):     return v.hour + v.minute / 60.0
    if isinstance(v, datetime):
        if v.year == 1900 and v.month == 1 and v.day == 1:
            return v.hour + v.minute / 60.0
        return 0.0
    t = _time_parse(v)
    if t: return t.hour + t.minute / 60.0
    return _float(v)


def _build_merged_lookup(ws):
    lookup = {}
    for mr in ws.merged_cells.ranges:
        av = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = av
    return lookup


def _cell(ws, r, c, lookup):
    v = ws.cell(r, c).value
    return v if v is not None else lookup.get((r, c))


def _strip_prefix(text: str, *prefixes: str) -> str:
    s = _clean(text)
    for p in prefixes:
        rx = rf"^\s*{re.escape(p)}\s*:?\s*"
        new = re.sub(rx, "", s, flags=re.IGNORECASE)
        if new != s: return new.strip()
    return s


def _extract_date_from_string(text: str):
    """Find a date like 02/05/2025 or 02-05-2025 in a string. Skip --/--/."""
    s = _clean(text)
    m = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", s)
    if not m: return None
    return _date_parse(f"{m.group(1)}/{m.group(2)}/{m.group(3)}")


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------
def parse_tp173(source: Union[Path, str, BytesIO]) -> dict:
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    ws = wb.active
    L = _build_merged_lookup(ws)

    # =====================================================================
    # HEADER (rows 4-7)
    # =====================================================================
    header = {}

    # Date — S4
    header["date"]       = _date_parse(_cell(ws, 4, 19, L))   # S4
    header["day_number"] = _int(_cell(ws, 2, 21, L))          # U2

    # Well — A6 (bare name, no "Well :" prefix)
    well_raw = _clean(_cell(ws, 6, 1, L) or "")
    header["well_name"]  = _strip_prefix(well_raw, "Well :", "Well", "Puits", "PUITS")

    # Rig — F6 "APPAREIL: TP-173"
    rig_raw = _clean(_cell(ws, 6, 6, L) or "")
    header["rig_name"]   = _strip_prefix(rig_raw, "APPAREIL", "RIG :", "RIG", "APPAREIL :")

    # Class / Zone — H6, J6
    cls = _clean(_cell(ws, 6, 8, L) or "")
    if cls: header["well_class"] = cls
    zone = _clean(_cell(ws, 6, 10, L) or "")
    if zone: header["field_name"] = zone

    # BOP tests — combined into K6 / N6 strings like "Last BOP's Test --/--/2026"
    bop_last = _cell(ws, 6, 11, L)
    if bop_last:
        d = _extract_date_from_string(str(bop_last))
        if d: header["bop_test"] = d
    bop_next = _cell(ws, 6, 14, L)
    if bop_next:
        d = _extract_date_from_string(str(bop_next))
        if d: header["next_bop_test"] = d

    # Last CSG SHOE — K7 "DERNIER TUBAGE: Csg 7\"shoe @ 850m  BC @725m et BP..."
    csg_raw = _cell(ws, 7, 11, L)
    if csg_raw:
        s = _strip_prefix(str(csg_raw), "DERNIER TUBAGE", "Last Csg", "Dernier Csg")
        if s:
            # Frontend "Last CSG SHOE" reads from header.top_shoe via insert
            header["top_shoe"] = s

    # Mud type — S6 "TYPE: Saumure"
    mud_raw = _cell(ws, 6, 19, L)
    mud_type = _strip_prefix(_clean(mud_raw or ""), "TYPE")

    # Well objective — A10/A11 merged
    obj = _cell(ws, 10, 1, L) or _cell(ws, 8, 1, L)
    if obj:
        s = _strip_prefix(str(obj), "BUT DU WORK OVER")
        if s: header["well_objective"] = s

    # =====================================================================
    # OPERATIONS  (rows 13-26; cols A=start, B=end, C=bill, D=description)
    # =====================================================================
    activities = []
    for row in range(13, 27):
        # Stop if we hit the "Remarques" header (D27)
        marker = _cell(ws, row, 4, L)
        if marker and "REMARQUE" in str(marker).upper():
            break
        start = _cell(ws, row, 1, L)   # A
        end   = _cell(ws, row, 2, L)   # B
        bill  = _clean(_cell(ws, row, 3, L) or "")   # C
        desc  = _clean(_cell(ws, row, 4, L) or "")   # D

        if not bill and not desc and start is None:
            continue
        # Real ops have a bill code (T1/T2/T3/T4/NR)
        if not bill:
            continue

        start_t = _time_parse(start)
        end_t   = _time_parse(end)
        hours = 0.0
        if start_t and end_t:
            sm = start_t.hour * 60 + start_t.minute
            em = end_t.hour * 60 + end_t.minute
            if em == sm: hours = 24.0
            elif em > sm: hours = (em - sm) / 60.0
            else: hours = (em + 1440 - sm) / 60.0

        activities.append({
            "start_time": start_t,
            "end_time":   end_t,
            "hours":      hours,
            "phase_name": "",
            "code": "", "sub": "",
            "description": desc,
            "start_md": 0, "end_md": 0,
            "npt": 0, "npt_detail": "",
            "npt_company": "", "op_company": "",
            "bill": bill,
        })

    # =====================================================================
    # TARIFF TOTALS  (rows 10-11 cols L/M/N/O/P/R)
    # H/jour at row 10, H. Cumul at row 11
    # =====================================================================
    tarif_totals = {}
    daily_t1 = _cell(ws, 10, 12, L)   # L10
    daily_t2 = _cell(ws, 10, 13, L)   # M10
    daily_t3 = _cell(ws, 10, 14, L)   # N10
    daily_t4 = _cell(ws, 10, 15, L)   # O10
    daily_nr = _cell(ws, 10, 16, L)   # P10
    for k, v in [("t1", daily_t1), ("t2", daily_t2), ("t3", daily_t3),
                 ("t4", daily_t4), ("nr", daily_nr)]:
        if v is not None:
            try: tarif_totals[k] = _duration_hours(v)
            except: pass

    cum_t1 = _cell(ws, 11, 12, L)
    cum_t2 = _cell(ws, 11, 13, L)
    cum_t3 = _cell(ws, 11, 14, L)
    cum_t4 = _cell(ws, 11, 15, L)
    cum_total = _cell(ws, 11, 18, L)
    for k, v in [("cum_t1", cum_t1), ("cum_t2", cum_t2), ("cum_t3", cum_t3),
                 ("cum_t4", cum_t4), ("cum_total", cum_total)]:
        if v is not None:
            try: tarif_totals[k] = _float(v)
            except: pass

    # =====================================================================
    # MUD CHECKS (rows 11-19 cols S label + T value)
    # =====================================================================
    mud_checks = {}
    if mud_type: mud_checks["mud_type"] = mud_type
    for row, key in [
        (11, "density"),    # Densité
        (12, "fun_vis"),    # Visc Mast
        (13, "apl_fl"),     # Filtrat
        (14, "yp"),         # YP@150°F
        (15, "hpht_fl"),    # HP/HT Filtrat
        (16, "es"),         # Elect stab
    ]:
        v = _cell(ws, row, 20, L)    # col T
        if v is None: continue
        try: mud_checks[key] = _float(v)
        except: pass

    # =====================================================================
    # MUD VOLUMES (rows 7-11 cols T value, with labels in S)
    # =====================================================================
    mud_volume = {}
    # Reception/Ejection at r7 T/V
    rec = _cell(ws, 7, 20, L)        # T7
    ej  = _cell(ws, 7, 22, L)        # V7
    if rec is not None and _clean(str(rec)) != "/":
        try: mud_volume["received_volume"] = _float(rec)
        except: pass
    if ej is not None:
        try: mud_volume["dumped_volume"] = _float(ej)
        except: pass
    # Perte surface T9, Perte P/T T10
    ps = _cell(ws, 9, 20, L)
    if ps is not None:
        try: mud_volume["surface_loss"] = _float(ps)
        except: pass
    pt = _cell(ws, 10, 20, L)
    if pt is not None:
        try: mud_volume["formation_loss"] = _float(pt)
        except: pass

    # =====================================================================
    # MUD CHEMICAL USAGE (rows 22-33 cols S item, T initial, U used, V final)
    # =====================================================================
    chemicals = []
    for row in range(22, 34):
        item = _clean(_cell(ws, row, 19, L) or "")     # S
        if not item or item.upper() in ("PRODUITS",):
            continue
        # Skip if it's the "Eau m³" final row
        initial = _cell(ws, row, 20, L)                # T
        used    = _cell(ws, row, 21, L)                # U
        final   = _cell(ws, row, 22, L)                # V
        if initial is None and used is None and final is None:
            continue
        chemicals.append({
            "item":     item,
            "units":    "",
            "received": _clean(str(initial) if initial is not None else ""),
            "used":     _clean(str(used)    if used    is not None else ""),
            "on_loc":   _clean(str(final)   if final   is not None else ""),
        })

    # =====================================================================
    # PERSONNEL  (rows 27-31 col K label + col R count) — counts only
    # =====================================================================
    personnel = []
    for row in range(27, 32):
        role  = _clean(_cell(ws, row, 11, L) or "")     # K
        count = _cell(ws, row, 18, L)                    # R
        if not role or "PERSONNEL" in role.upper():
            continue
        # Skip the "Total" summary row
        if role.upper().startswith("TOTAL"):
            continue
        personnel.append({
            "company": role,
            "number":  _int(count, 0),
            "hours":   "",
            "names":   "",
        })

    # =====================================================================
    # SUPERVISOR — S35: "Représentant maître d'œuvre :M,EL BERRICHI"
    # The label and name are combined into a single string.  Per the
    # two-supervisor rule, this one name goes to header.supervisor (only
    # one name on this template, so superintendent stays empty).
    # =====================================================================
    sup_raw = _cell(ws, 35, 19, L)
    if sup_raw:
        name = _strip_prefix(str(sup_raw), "Représentant maître d'œuvre",
                             "Représentant maître d'oeuvre",
                             "Représentant maitre d'œuvre",
                             "Représentant maitre d'oeuvre")
        # Strip a stray leading colon (we removed the label but maybe not the ":")
        name = name.lstrip(": ").strip()
        if name:
            header["supervisor"] = name

    # =====================================================================
    # TEXT SECTIONS
    #   r31 / r32: SITUATION AU RAPPORT in col A, value in col F
    #   r33 / r34: PROGRAMME PREVU in col A, value in col F
    #   r27 / r28: Remarques (D27 label, D28 value)
    # =====================================================================
    text_sections = {}

    # Situation
    sit_label = _cell(ws, 31, 1, L) or _cell(ws, 32, 1, L) or ""
    if "SITUATION" in str(sit_label).upper():
        for r in (31, 32):
            v = _cell(ws, r, 6, L)
            if v:
                vv = _clean(v)
                text_sections["current_operation"] = vv
                text_sections["day_summary"]       = vv
                break

    # Programme prévu
    plan_label = _cell(ws, 33, 1, L) or _cell(ws, 34, 1, L) or ""
    if "PROGRAMME" in str(plan_label).upper():
        for r in (33, 34):
            v = _cell(ws, r, 6, L)
            if v:
                text_sections["plan_operations"] = _clean(v)
                break

    # Remarques (D27 label, D28+ value)
    rem_label = _cell(ws, 27, 4, L)
    if rem_label and "REMARQUE" in str(rem_label).upper():
        v = _cell(ws, 28, 4, L)
        if v:
            text_sections["remarks"] = _clean(v)

    # =====================================================================
    # SAFETY — no accident-free counter on this template
    # =====================================================================
    safety = {}

    wb.close()

    return {
        "header": header,
        "activities": activities,
        "text_sections": text_sections,
        "mud_checks": mud_checks,
        "mud_volume": mud_volume,
        "mud_chemical_usage": chemicals,
        "personnel_data": personnel,
        "pumps": [],
        "well_location": {},
        "survey_data": [],
        "safety": safety,
        "tarif_totals": tarif_totals,
    }


# Drop-in compat
parse_daily_excel_report = parse_tp173


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Usage: tp173_extract.py SOURCE.xlsx")
    data = parse_tp173(Path(sys.argv[1]))

    def default(o):
        if isinstance(o, (date_type, datetime)): return o.isoformat()
        if isinstance(o, time): return o.strftime("%H:%M:%S")
        if isinstance(o, timedelta): return o.total_seconds()
        return str(o)
    print(json.dumps(data, indent=2, default=default, ensure_ascii=False))
