#!/usr/bin/env python3
"""
enf04_extract.py — extract an ENF#04 (ENAFOR rig 04, Haoud Berkaoui, DDNH
wells) Daily Workover Report into the standard dict shape.

Source layout
-------------
Compact French workover template (~71 rows × 20 cols).  Title at B9:
"RAPPORT JOURNALIER DE WORKOVER" (note: "DE" not "DU"  — distinguishes it
from the TP-179 format).  Regional direction at B7: "HAOUD BERKAOUI".
Sheet name: "Rapport SHDP".

Header at rows 10-14 (label/value pairs in distinct cells).
Operations at rows 16-19 (Debut/Fin/Temp H/Chronologie).
After-midnight section starts at r34 (often empty).
Current operation @ 6h00 at r50.
Mud panel at rows 57-61 (compact: viscosity, density, gel, Pv, YP +
volume in surface, loss, total, water).
Personnel at r70 (single row with counts only: Maitre d'œuvre + Entrepreneur
+ Total).  No individual supervisor names on this template.
"""
from __future__ import annotations
import re
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers (same shape as the other extractors)
# ---------------------------------------------------------------------------
def _clean(v) -> str:
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _float(v, default=0.0) -> float:
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    s = _clean(v).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "None"): return default
    s = re.sub(r"[a-zA-Zé°%/³]+$", "", s).strip()
    try: return float(s)
    except ValueError: return default


def _int(v, default=0) -> int:
    return int(_float(v, float(default)))


def _date_parse(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date_type): return v
    s = _clean(v)
    # Strip "du:" prefix if present (e.g. "du: 10/05/2026")
    s = re.sub(r"^du\s*:?\s*", "", s, flags=re.IGNORECASE).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None


def _time_parse(v):
    if v is None: return None
    if isinstance(v, time): return v
    if isinstance(v, datetime):
        # Excel epoch zero (1900-01-01) means 24:00 / end of day
        return v.time()
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        if total == 86400: return time(0, 0)
        return time((total // 3600) % 24, (total % 3600) // 60)
    s = _clean(v)
    if not s or s in ("-", "None"): return None
    if s in ("24:00", "24:00:00"): return time(0, 0)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try: return datetime.strptime(s, fmt).time()
        except ValueError: continue
    return None


def _duration_hours(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, timedelta): return v.total_seconds() / 3600.0
    if isinstance(v, time):     return v.hour + v.minute / 60.0
    if isinstance(v, datetime):
        if v.year == 1900 and v.month == 1 and v.day == 1:
            return v.hour + v.minute / 60.0
        return 0.0
    t = _time_parse(v)
    if t: return t.hour + t.minute / 60.0
    return _float(v)


def _build_merged_lookup(ws):
    lookup = {}
    for mr in ws.merged_cells.ranges:
        av = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = av
    return lookup


def _cell(ws, r, c, lookup):
    v = ws.cell(r, c).value
    return v if v is not None else lookup.get((r, c))


def _normalize_rig(name: str) -> str:
    """'ENAFOR # 04' or 'ENAFOR #04' or 'ENAFOR # 4' -> 'ENF#04'."""
    s = _clean(name).upper()
    m = re.search(r"ENAFOR\s*#?\s*(\d+)", s)
    if m:
        return f"ENF#{int(m.group(1)):02d}"
    return _clean(name)


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------
def parse_enf04(source: Union[Path, str, BytesIO]) -> dict:
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    ws = wb.active
    L = _build_merged_lookup(ws)

    # =====================================================================
    # HEADER (rows 10-14)
    # =====================================================================
    header = {}

    # Row 10: date + day number
    header["date"]       = _date_parse(_cell(ws, 10, 5,  L))   # E10
    header["day_number"] = _int(_cell(ws, 10, 15, L))          # O10

    # Row 11: well / nature / rig
    header["well_name"]  = _clean(_cell(ws, 11, 6,  L))        # F11
    header["well_type"]  = _clean(_cell(ws, 11, 12, L))        # L11 (PPH)
    header["rig_name"]   = _normalize_rig(_cell(ws, 11, 19, L) or "")  # S11

    # Row 12: Last CSG (Dernier CSG) — size + top + shoe
    csg_size  = _clean(_cell(ws, 12, 6,  L) or "")             # F12
    csg_top   = _clean(_cell(ws, 12, 12, L) or "")             # L12
    csg_shoe  = _cell(ws, 12, 19, L)                            # S12
    if csg_size or csg_shoe is not None:
        # Frontend "Last CSG SHOE" reads from DB.lastCSNlnrSHOE, which is filled
        # from header.top_shoe.  Format as 'SIZE @ DEPTHm'.
        if csg_shoe is not None:
            try:
                depth = int(float(str(csg_shoe)))
                header["top_shoe"] = f"{csg_size} @ {depth}m" if csg_size else f"@ {depth}m"
            except (ValueError, TypeError):
                header["top_shoe"] = f"{csg_size} @ {_clean(csg_shoe)}" if csg_size else _clean(csg_shoe)
        else:
            header["top_shoe"] = csg_size
    # Last CSG Top (rarely meaningful — usually "0 m") — only set if non-zero
    if csg_top and csg_top.replace(" m", "").strip() not in ("", "0", "0m"):
        header["last_csg_top"] = csg_top

    # Row 13: Volume du puits + Hauteur/Bottom Perfos
    well_vol = _cell(ws, 13, 6,  L)                            # F13
    h_perfos = _cell(ws, 13, 12, L)                            # L13
    b_perfos = _cell(ws, 13, 19, L)                            # S13
    if well_vol is not None: header["well_volume"] = _float(well_vol)
    if h_perfos is not None: header["h_perfos"]    = _float(h_perfos)
    if b_perfos is not None: header["b_perfos"]    = _float(b_perfos)

    # Row 14: Objective
    obj = _cell(ws, 14, 6, L)                                  # F14
    if obj:
        header["well_objective"] = _clean(obj)

    # =====================================================================
    # OPERATIONS  (rows 16-19; header at r15: Debut/Fin/Temp H/Chronologie)
    # =====================================================================
    activities = []
    for row in range(16, 34):     # generous range; stops at "Après Minuit" label
        start = _cell(ws, row, 2, L)        # B
        end   = _cell(ws, row, 3, L)        # C
        hrs   = _cell(ws, row, 4, L)        # D
        desc  = _clean(_cell(ws, row, 5, L) or "")     # E

        # Stop at the "Après Minuit" / "Situation à 6H00" labels
        if any(t in desc.upper() for t in ("APRÈS MINUIT", "SITUATION", "PRESSION:",
                                            "OUTILS")):
            break
        if not desc and start is None:
            continue

        start_t = _time_parse(start)
        end_t   = _time_parse(end)
        hours = _duration_hours(hrs)
        if hours == 0.0 and start_t and end_t:
            sm = start_t.hour * 60 + start_t.minute
            em = end_t.hour * 60 + end_t.minute
            if em == sm:    hours = 24.0
            elif em > sm:   hours = (em - sm) / 60.0
            else:           hours = (em + 1440 - sm) / 60.0

        activities.append({
            "start_time": start_t,
            "end_time":   end_t,
            "hours":      hours,
            "phase_name": "",
            "code": "", "sub": "",
            "description": desc,
            "start_md": 0, "end_md": 0,
            "npt": 0, "npt_detail": "",
            "npt_company": "", "op_company": "",
            "bill": "",        # this template doesn't tag per-row bill codes
        })

    # =====================================================================
    # AFTER MIDNIGHT  (rows 35-48 — usually empty on this template)
    # =====================================================================
    am_lines = []
    for row in range(35, 49):
        start = _cell(ws, row, 2, L)
        end   = _cell(ws, row, 3, L)
        desc  = _clean(_cell(ws, row, 5, L) or "")
        if not desc: continue
        if "PRESSION" in desc.upper() or "SITUATION" in desc.upper(): break
        am_lines.append(f"{_clean(start) or ''} - {_clean(end) or ''} : {desc}")

    # =====================================================================
    # TEXT SECTIONS  (rows 49-56)
    # =====================================================================
    text_sections = {}

    # Pressure (B49)
    pressure = _cell(ws, 49, 2, L)
    if pressure:
        header["pressure_info"] = _clean(pressure)

    # Current operation @ 6h00 (B50 = label "Situation à 6H00", E50 = value)
    sit_label = _clean(_cell(ws, 50, 2, L) or "")
    if "SITUATION" in sit_label.upper():
        sit_val = _cell(ws, 50, 5, L)
        if sit_val:
            v = _clean(sit_val)
            text_sections["current_operation"] = v
            text_sections["day_summary"]       = v

    # After-midnight text section
    if am_lines:
        text_sections["after_midnight"] = " | ".join(am_lines)

    # Well-under-pressure / casing description (r56 B starts with "Puite sous: BSS")
    well_sous = _cell(ws, 56, 2, L)
    if well_sous and "BSS" in str(well_sous).upper():
        header["well_under"] = _clean(well_sous)

    # =====================================================================
    # MUD CHECKS  (rows 57-61, cols B label + D value, left side)
    # =====================================================================
    mud_checks = {}
    fv = _cell(ws, 57, 4, L)         # D57 - Fv (funnel viscosity)
    if fv is not None: mud_checks["fun_vis"] = _float(fv)
    d  = _cell(ws, 58, 4, L)         # D58 - Densité
    if d is not None:  mud_checks["density"] = _float(d)
    gel = _cell(ws, 59, 4, L)        # D59 - Gel 0/10
    if gel is not None:
        s = _clean(gel)
        if "/" in s:
            try:
                a, b = s.split("/", 1)
                mud_checks["gel10sec"] = _float(a); mud_checks["gel10m"] = _float(b)
            except ValueError:
                pass
        else:
            mud_checks["gel10sec"] = _float(gel)
    pv = _cell(ws, 60, 4, L)         # D60 - Pv
    if pv is not None: mud_checks["pv"] = _float(pv)
    yp = _cell(ws, 61, 4, L)         # D61 - YP
    if yp is not None: mud_checks["yp"] = _float(yp)

    # =====================================================================
    # MUD VOLUMES  (rows 57-61, col H label + col L value)
    # =====================================================================
    mud_volume = {}
    pits = _cell(ws, 57, 12, L)      # L57 - Volume en Surface
    if pits is not None: mud_volume["pits_volume"] = _float(pits)
    loss = _cell(ws, 58, 12, L)      # L58 - Perte accidentelle/Tripping
    if loss is not None: mud_volume["surface_loss"] = _float(loss)
    fmt_loss = _cell(ws, 59, 12, L)  # L59 - Perte formation
    if fmt_loss is not None: mud_volume["formation_loss"] = _float(fmt_loss)
    total = _cell(ws, 60, 12, L)     # L60 - Volume total
    if total is not None: mud_volume["total_volume"] = _float(total)
    water = _cell(ws, 61, 12, L)     # L61 - Volume d'eau
    if water is not None: mud_volume["water_volume"] = _float(water)
    # If total wasn't given, compute it
    if "total_volume" not in mud_volume and mud_volume:
        mud_volume["total_volume"] = sum(
            v for k, v in mud_volume.items() if k == "pits_volume" and v is not None
        )

    # =====================================================================
    # TARIFF TOTALS  (rows 64-67: T1/T2/T3/NR hours)
    # =====================================================================
    tarif_totals = {}
    for row, bill_col_letter, bill_key in [
        (64, "G", "T1"), (65, "G", "T2"), (66, "G", "T3"), (67, "G", "NR"),
    ]:
        code = _clean(_cell(ws, row, 7, L) or "")            # col G
        hrs  = _cell(ws, row, 8, L)                           # col H
        if code == bill_key and hrs is not None:
            try:
                tarif_totals[bill_key.lower()] = _duration_hours(hrs)
            except Exception:
                pass

    # =====================================================================
    # PERSONNEL  (row 70 — single row with counts only, no individual names)
    # =====================================================================
    personnel = []
    moe = _cell(ws, 70, 5, L)             # E70 = Maitre d'œuvre count
    ent = _cell(ws, 70, 15, L)            # O70 = Entrepreneur count
    tot = _cell(ws, 70, 19, L)            # S70 = Total count
    if moe is not None:
        personnel.append({"company": "Maitre d'œuvre", "number": _int(moe, 0),
                          "hours": "", "names": ""})
    if ent is not None:
        personnel.append({"company": "Entrepreneur", "number": _int(ent, 0),
                          "hours": "", "names": ""})

    # No individual supervisor names on this template, so supervisor /
    # superintendent stay unset (per the two-supervisor convention).

    # =====================================================================
    # Back-assign the bill code(s) to each operation row.
    #
    # This template doesn't tag individual operations with T1/T2/T3/NR —
    # only the daily totals at rows 64-67 know the breakdown.  We partition
    # the ops into groups whose hours sum exactly to each T-bucket total.
    # See bill_code_assign.assign_bill_codes for the algorithm.
    # =====================================================================
    from helpers.bill_code_assign import assign_bill_codes
    activities = assign_bill_codes(activities, tarif_totals)

    wb.close()

    return {
        "header": header,
        "activities": activities,
        "text_sections": text_sections,
        "mud_checks": mud_checks,
        "mud_volume": mud_volume,
        "mud_chemical_usage": [],
        "personnel_data": personnel,
        "pumps": [],
        "well_location": {},
        "survey_data": [],
        "safety": {},
        "tarif_totals": tarif_totals,
    }


# Drop-in compat
parse_daily_excel_report = parse_enf04


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Usage: enf04_extract.py SOURCE.xlsx")
    data = parse_enf04(Path(sys.argv[1]))

    def default(o):
        if isinstance(o, (date_type, datetime)): return o.isoformat()
        if isinstance(o, time): return o.strftime("%H:%M:%S")
        if isinstance(o, timedelta): return o.total_seconds()
        return str(o)
    print(json.dumps(data, indent=2, default=default, ensure_ascii=False))