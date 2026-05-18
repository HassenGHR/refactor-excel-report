"""
ddr_extract.py — Extract structured data from an ENAFOR Daily Drilling Report.

Replaces the template-based conversion in ddr_to_workover.py with a clean
data-only extraction. Output is a dict that matches the format expected by
insert_parsed_report() — the same structure that
`excel_import_service.parse_daily_excel_report()` returns.

CLI:
    python ddr_extract.py SOURCE.xlsx [-o OUTPUT.json] [--pretty]

The script can also be imported and called programmatically:

    from ddr_extract import parse_ddr
    data = parse_ddr(Path("report.xlsx"))
    # data["header"]["well_name"]  -> 'TAOP#06'
    # data["activities"][0]["start_time"] -> datetime.time(0, 0)
    # data["mud_checks"]["density"] -> 1.37
"""

from __future__ import annotations
import argparse
import json
import re
import sys
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


# ===========================================================================
# Generic helpers (compatible with excel_import_service.py)
# ===========================================================================

def _clean(text) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def _float(text, default: float = 0.0) -> float:
    if text is None:
        return default
    if isinstance(text, (int, float)):
        return float(text)
    s = _clean(str(text)).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "()", "None"):
        return default
    s = re.sub(r"[a-zA-Z%]+$", "", s).strip()
    if not s:
        return default
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def _int(text, default: int = 0) -> int:
    return int(_float(text, float(default)))


def _date_parse(val) -> Optional[date_type]:
    """Parse a date from many representations: datetime, date, 'dd/mm/yyyy',
    'dd-mm-yyyy', and the ENAFOR-specific 'dd,mm,yy' (e.g. '24,04,26')."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date_type):
        return val
    text = _clean(str(val))
    # ENAFOR BOP test format: dd,mm,yy
    m = re.match(r"^\s*(\d{1,2})[,\.\-/](\d{1,2})[,\.\-/](\d{2,4})\s*$", text)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        if y < 100:
            y += 2000
        try:
            return date_type(y, mo, d)
        except ValueError:
            pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d %m %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _time_from_cell(val) -> Optional[time]:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, timedelta):
        secs = int(val.total_seconds())
        if secs < 0:
            return None
        return time((secs // 3600) % 24, (secs % 3600) // 60)
    text = _clean(str(val))
    if text in ("", "-", "None"):
        return None
    if text == "24:00":
        return time(0, 0)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M", "%H.%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def _hours_from_cell(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, timedelta):
        return val.total_seconds() / 3600
    if isinstance(val, time):
        return val.hour + val.minute / 60.0
    if isinstance(val, datetime):
        return val.hour + val.minute / 60.0
    return _float(val)


def _build_merged_lookup(ws) -> Dict[Tuple[int, int], Any]:
    lookup: Dict[Tuple[int, int], Any] = {}
    for mr in ws.merged_cells.ranges:
        val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if r != mr.min_row or c != mr.min_col:
                    lookup[(r, c)] = val
    return lookup


def _cell(ws, row: int, col: int, lookup: Dict) -> Any:
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return val
    return lookup.get((row, col))


def _scan_for(ws, lookup, keyword, row_range, col_range):
    kw = keyword.upper()
    for r in range(row_range[0], row_range[1] + 1):
        for c in range(col_range[0], col_range[1] + 1):
            v = _cell(ws, r, c, lookup)
            if v and kw in str(v).upper():
                return (r, c)
    return None


# Unit conversion
BBL_TO_M3 = 0.158987294928


# ===========================================================================
# ENAFOR DDR-specific extraction
# ===========================================================================

def _extract_header(ws, lookup) -> Dict[str, Any]:
    """Header section: rig, well, field, date, day, depths, BHA, signatures."""
    h: Dict[str, Any] = {}

    # Rig name: A3 (e.g. 'ENF # 17' → 'ENF#17')
    rig = _cell(ws, 3, 1, lookup)
    if rig:
        h["rig_name"] = re.sub(r"\s*#\s*", "#", _clean(str(rig)))

    # Report date: W4 (col 23)
    d = _date_parse(_cell(ws, 4, 23, lookup))
    if d:
        h["date"] = d

    # Field: G5
    field = _cell(ws, 5, 7, lookup)
    if field:
        h["field_name"] = _clean(str(field))

    # Well: G6
    well = _cell(ws, 6, 7, lookup)
    if well:
        h["well_name"] = _clean(str(well)).replace("# ", "#").replace(" #", "#")

    # Previous depth: M6 (text like '4900m' — keep as string + numeric)
    pdepth = _cell(ws, 6, 13, lookup)
    if pdepth is not None:
        h["previous_depth_text"] = _clean(str(pdepth))
        h["tmd"] = _float(pdepth)

    # Present depth: J5 (col 10, row 5)  — usually present
    present = _cell(ws, 5, 10, lookup)
    if present is not None and _float(present) > 0:
        h["present_depth"] = _float(present)

    # Drill time (h): S6 — interpret as drill-time hours, not day count
    dt = _cell(ws, 6, 19, lookup)
    if dt is not None:
        h["drill_time_hours"] = _float(dt)

    # Report N°: scan a small area for an "N°" label and read the value
    # adjacent to it. Skip datetimes (they leak in via merge propagation).
    pos = _scan_for(ws, lookup, "N°", (1, 8), (20, 28))
    if pos:
        for dc in (1, 2, 3):
            n = _cell(ws, pos[0], pos[1] + dc, lookup)
            if n is None or isinstance(n, (datetime, date_type, time)):
                continue
            txt = _clean(str(n))
            if txt in ("", "N°"):
                continue
            digits = re.sub(r"\D", "", txt)
            if digits and len(digits) < 6:           # sanity-cap (a report N° is small)
                h["day_number"] = int(digits)
                break

    # BOP test date: U9 (format 'dd,mm,yy')
    bop = _date_parse(_cell(ws, 9, 21, lookup))
    if bop:
        h["bop_test"] = bop

    # Both names on the ENAFOR report are 12-hour-shift supervisors. The DB
    # has separate `supervisor` and `superintendent` columns, so we use them
    # as slot 1 and slot 2 — the second supervisor lands in `superintendent`
    # even though the column name is technically misleading.
    # TP Sénior (G3) → supervisor; TP Junior (G4) → superintendent (= second supervisor)
    ts = _cell(ws, 3, 7, lookup)
    if ts:
        h["supervisor"] = _clean(str(ts))
    tj = _cell(ws, 4, 7, lookup)
    if tj:
        h["superintendent"] = _clean(str(tj))

    # Last casing: D11 (size) + E11 (depth)
    csg_size = _cell(ws, 11, 4, lookup)
    csg_depth = _cell(ws, 11, 5, lookup)
    if csg_size:
        h["last_csg_size"] = _clean(str(csg_size))
        if csg_depth:
            h["last_csg_depth"] = _float(csg_depth)
            h["last_csg_shoe"] = f'{_clean(str(csg_size))} @ {csg_depth}m'

    # Top shoe: D12 (size) + E12 (depth)
    ts_size = _cell(ws, 12, 4, lookup)
    ts_depth = _cell(ws, 12, 5, lookup)
    if ts_size:
        h["top_shoe_size"] = _clean(str(ts_size))
        if ts_depth:
            h["top_shoe_depth"] = _float(ts_depth)
            h["top_shoe"] = f'{_clean(str(ts_size))} @ {ts_depth}m'

    # Top liner: D13 (size only — depth often missing)
    tl_size = _cell(ws, 13, 4, lookup)
    if tl_size:
        h["top_liner_size"] = _clean(str(tl_size))
        h["last_csg_top"] = _clean(str(tl_size))

    # BHA: G11 (length) + J11 (details)
    bha_len = _cell(ws, 11, 7, lookup)
    if bha_len:
        h["bha_length"] = _clean(str(bha_len))
    bha_det = _cell(ws, 11, 10, lookup)
    if bha_det:
        h["bha_details"] = _clean(str(bha_det))

    return h


def _extract_activities(ws, lookup) -> List[Dict[str, Any]]:
    """Operations / timing block. Walks rows starting at 20 in cols A..E.
    Continuation rows (empty start/end but non-empty description) are
    appended to the previous operation's description.

    The block doesn't have a fixed end row — different reports have
    different numbers of operations.  We walk until we hit a section
    boundary (the "AFTER MIDNIGHT" / "PLAN OPERATION" / "REMARKS" labels
    that follow the ops table) or until we've scanned 30 rows past the
    start (a safety cap — no real report has that many ops).
    """
    ops: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None

    # Stop markers that signal the end of the operations block
    STOP_MARKERS = ("AFTER MIDNIGHT", "APRÈS MINUIT", "APRES MINUIT",
                    "PLAN OPER", "PROGRAMME", "SITUATION", "REMARKS",
                    "REMARQUES", "PERSONNEL", "MUD CHECK")

    OPS_START = 20
    OPS_MAX_END = 50      # generous cap; we break out earlier on stop marker

    for r in range(OPS_START, OPS_MAX_END):
        start_val = _cell(ws, r, 1, lookup)
        end_val   = _cell(ws, r, 2, lookup)
        hours_val = _cell(ws, r, 3, lookup)
        code_val  = _cell(ws, r, 4, lookup)
        desc_val  = _cell(ws, r, 5, lookup)

        # Check for stop-marker text in col A (or anywhere on the row's first
        # few cells).  When found, finalize the current op and exit.
        row_marker_text = " ".join(
            str(_cell(ws, r, c, lookup) or "") for c in range(1, 6)
        ).upper()
        if any(m in row_marker_text for m in STOP_MARKERS):
            break

        start_t = _time_from_cell(start_val)
        end_t   = _time_from_cell(end_val)

        if start_t is not None and end_t is not None:
            if current:
                ops.append(current)
            current = {
                "start_time":   start_t,
                "end_time":     end_t,
                "hours":        _hours_from_cell(hours_val),
                "phase_name":   "",
                "code":         "",
                "sub":          "",
                "description":  _clean(str(desc_val or "")),
                "start_md":     0,
                "end_md":       0,
                "npt":          0,
                "npt_detail":   "",
                "npt_company":  "",
                "op_company":   "",
                "bill":         _clean(str(code_val or "")),
            }
        elif desc_val and current is not None:
            extra = _clean(str(desc_val))
            current["description"] = (current["description"] + "\n" + extra).strip()
    if current:
        ops.append(current)
    return ops


def _extract_text_sections(ws, lookup, activities) -> Dict[str, str]:
    """Narrative text blocks."""
    sec: Dict[str, str] = {}

    # After midnight — scan for "AFTER MIDNIGHT" / "APRÈS MINUIT" label;
    # the value sits on the same row in col E, or on the row below.
    # Different reports have different numbers of operations so the label
    # row isn't fixed.
    pos = _scan_for(ws, lookup, "AFTER MIDNIGHT", (30, 55), (1, 8))
    if not pos:
        pos = _scan_for(ws, lookup, "APRÈS MINUIT", (30, 55), (1, 8))
    if not pos:
        pos = _scan_for(ws, lookup, "APRES MINUIT", (30, 55), (1, 8))
    if pos:
        # Try same row col E first, then col E one row down
        for r_off, c in [(0, 5), (1, 5), (0, 3), (1, 3)]:
            v = _cell(ws, pos[0] + r_off, c, lookup)
            if v:
                text = _clean(str(v))
                # If we picked up the label itself, strip it
                text = re.sub(r"^\s*(?:AFTER\s*MIDNIGHT|APR[ÈE]S\s*MINUIT)\s*:?\s*",
                              "", text, flags=re.IGNORECASE)
                if text and "AFTER MIDNIGHT" not in text.upper() and "MINUIT" not in text.upper():
                    sec["after_midnight"] = text
                    break

    # Plan operations: scan for "PLAN" (handles "Plan Opetations" typo too)
    pos = _scan_for(ws, lookup, "PLAN", (40, 70), (1, 12))
    if pos:
        # Value usually one column to the right of label
        for dc in (1, 2, 3):
            v = _cell(ws, pos[0], pos[1] + dc, lookup)
            if v and "PLAN" not in str(v).upper():
                sec["plan_operations"] = _clean(str(v))
                break

    # "Situation au Rapport" — best inferred from operations: the last one
    # whose description contains "IN PROGRESS".
    # Capped at 300 chars because the frontend "Situation" field has a
    # display limit (db column is text but the UI truncates / overflows
    # past ~300 chars).
    SITUATION_MAX = 300
    if activities:
        in_prog = [o for o in activities if "IN PROGRESS" in (o.get("description") or "").upper()]
        chosen = in_prog[-1] if in_prog else activities[-1]
        text = chosen["description"] or ""
        if len(text) > SITUATION_MAX:
            # Cut at the last whitespace before the limit so we don't break
            # mid-word.  Append an ellipsis to flag the truncation.
            cut = text[:SITUATION_MAX].rsplit(None, 1)[0]
            text = cut + "…"
        sec["current_operation"] = text
        sec["day_summary"] = text

    return sec


def _extract_mud_checks(ws, lookup) -> Dict[str, Any]:
    """Mud properties from the Pumps / WOB / Mud panel (rows ~24-30, cols R-U)."""
    mud: Dict[str, Any] = {}

    def _put(key, row, col):
        v = _cell(ws, row, col, lookup)
        if v is not None and (isinstance(v, (int, float)) or _clean(str(v)) not in ("", "-")):
            mud[key] = _float(v) if not isinstance(v, str) or re.search(r"\d", v) else _clean(str(v))

    # Density: S26 (col 19)
    _put("density",   26, 19)
    # Viscosity (V.Marsh): U26 (col 21)
    _put("fun_vis",   26, 21)
    # Solide %: S27 — whole-number percent (e.g. 30 means 30%)
    v = _cell(ws, 27, 19, lookup)
    if v is not None:
        mud["solid"] = _float(v)
    # Yield point: U27
    _put("yp",        27, 21)
    # Gel 10": S28
    _put("gel10sec",  28, 19)
    # Water: U29
    _put("h2o",       29, 21)
    # LGS: U30
    _put("lgs",       30, 21)

    # Oil/Water ratio: S30 — text like "86/14"
    ow = _cell(ws, 30, 19, lookup)
    if ow:
        mud["oil_water_ratio"] = _clean(str(ow))

    # Mud type — implicit (OBM is standard for ENAFOR; not always cell-tagged).
    # Look for a "Type" label near the header area.
    pos = _scan_for(ws, lookup, "MUD", (5, 12), (15, 25))
    if pos:
        v = _cell(ws, pos[0], pos[1] + 1, lookup)
        if v:
            mud["mud_type"] = _clean(str(v))

    return mud


def _extract_mud_volume(ws, lookup) -> Dict[str, Any]:
    """Hole volume and total circulation volume — source in bbl, converted to m³."""
    vol: Dict[str, Any] = {}

    hole_bbl = _cell(ws, 24, 19, lookup)        # S24
    if hole_bbl is not None:
        vol["string_volume_bbl"] = _float(hole_bbl)
        vol["string_volume"]     = round(_float(hole_bbl) * BBL_TO_M3, 2)

    circ_bbl = _cell(ws, 24, 21, lookup)        # U24
    if circ_bbl is not None:
        vol["total_volume_bbl"]  = _float(circ_bbl)
        vol["total_volume"]      = round(_float(circ_bbl) * BBL_TO_M3, 2)

    # Rig water (m³): T35
    rw = _cell(ws, 35, 20, lookup)
    if rw is not None:
        vol["rig_water_m3"] = _float(rw)

    return vol


def _extract_personnel(ws, lookup) -> List[Dict[str, Any]]:
    """Crew counts from rows 15 and 17 (Sup int / Elect / Mechanic etc.)."""
    rows = []
    # Categories and the (row, col) where the count lives. Labels on row 14
    # or row 16; values on row 15 or row 17 respectively.
    mapping = [
        ("S/Tool Pusher",            15, 3),
        ("J/Tool Pusher",            15, 5),
        ("Driller",                  15, 7),
        ("A/Driller",                15, 10),
        ("Cont. Lab",                15, 12),
        ("Cas. Lab",                 15, 15),
        ("Militaires/Vigiles",       15, 17),
        ("SPV (Elec/Mec/TDS)",       15, 18),
        ("DTR/Personnels",           15, 19),
        ("Front loader/Cariste",     15, 21),
        ("Intendant",                15, 25),
        ("Elect",                    17, 1),
        ("Mechanic",                 17, 3),
        ("Driver",                   17, 5),
        ("Catering",                 17, 7),
        ("Company man",              17, 9),
        ("Trainee",                  17, 12),
        ("HSE",                      17, 15),
        ("ADG",                      17, 17),
        ("Medec/Infirmier",          17, 18),
        ("Welder",                   17, 19),
        ("Other",                    17, 21),
    ]
    for label, r, c in mapping:
        v = _cell(ws, r, c, lookup)
        if v is None or _clean(str(v)) == "":
            continue
        text = _clean(str(v))
        # Parse counts: pure number, "X+Y" (sum), or text fallback
        if re.match(r"^\d+(\.\d+)?$", text):
            n = _int(text)
        elif re.match(r"^[\d.]+(\s*\+\s*[\d.]+)+$", text):
            n = sum(_int(p.strip()) for p in text.split("+"))
        elif re.match(r"^[\d.]+(/[\d.]+)+$", text):
            n = sum(_int(p.strip()) for p in text.split("/"))
        else:
            n = 0  # leave as 0; the raw string is preserved in "names"
        # Only keep "names" if it isn't just a plain number duplicate
        names_field = "" if re.match(r"^\d+(\.\d+)?$", text) else text
        rows.append({"company": label, "number": n, "hours": "", "names": names_field})
    return rows


def _extract_safety(ws, lookup) -> Dict[str, Any]:
    """HSE counters at the top of the report."""
    safety: Dict[str, Any] = {}
    # Accident-free days: F7
    afd = _cell(ws, 7, 6, lookup)
    if afd is not None:
        safety["accident_free_days"] = _int(afd)
    # HSE meetings: U7
    m = _cell(ws, 7, 21, lookup)
    if m is not None:
        safety["hse_meetings"] = _clean(str(m))
    # Permit to work: U8
    p = _cell(ws, 8, 21, lookup)
    if p is not None:
        safety["permits_to_work"] = _clean(str(p))
    # Stop card / stop tours: Y7
    s = _cell(ws, 7, 25, lookup)
    if s is not None:
        safety["stop_cards"] = _clean(str(s))
    # Exercice: Y8
    e = _cell(ws, 8, 25, lookup)
    if e is not None:
        safety["exercises"] = _clean(str(e))
    return safety


def _extract_tarif_totals(activities) -> Dict[str, float]:
    """Sum decimal hours per Tarif code (T1, T2, T3, T4, NR, …)."""
    tally: Dict[str, float] = {}
    for op in activities:
        code = (op.get("bill") or "").strip().upper()
        if code:
            tally[code] = tally.get(code, 0.0) + _hours_from_cell(op.get("hours"))
    return tally


# ===========================================================================
# Main entry point
# ===========================================================================

def parse_ddr(source) -> Dict[str, Any]:
    """Parse an ENAFOR Daily Drilling Report and return a structured dict.

    `source` may be a `Path`, a string path, or a file-like object (e.g. BytesIO).

    Output shape (compatible with insert_parsed_report()):
        {
          "header":           {...},
          "activities":       [{...}, ...],
          "text_sections":    {...},
          "mud_checks":       {...},
          "mud_volume":       {...},
          "mud_chemical_usage": [],   # not present in ENAFOR DDR
          "personnel_data":   [{...}, ...],
          "pumps":            [],     # placeholder
          "well_location":    {},     # placeholder
          "survey_data":      [],     # placeholder
          "safety":           {...},  # extra: HSE counters
          "tarif_totals":     {...},  # extra: T1/T2/T3/T4 daily totals
        }
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.active
    lookup = _build_merged_lookup(ws)

    header     = _extract_header(ws, lookup)
    activities = _extract_activities(ws, lookup)
    text_sec   = _extract_text_sections(ws, lookup, activities)

    result: Dict[str, Any] = {
        "header":             header,
        "activities":         activities,
        "text_sections":      text_sec,
        "mud_checks":         _extract_mud_checks(ws, lookup),
        "mud_volume":         _extract_mud_volume(ws, lookup),
        "mud_chemical_usage": [],
        "personnel_data":     _extract_personnel(ws, lookup),
        "pumps":              [],
        "well_location":      {},
        "survey_data":        [],
        # ENAFOR-DDR-specific extras (won't break insert_parsed_report)
        "safety":             _extract_safety(ws, lookup),
        "tarif_totals":       _extract_tarif_totals(activities),
    }

    wb.close()
    return result


def parse_daily_excel_report(file_bytes: bytes) -> Dict[str, Any]:
    """Drop-in compatibility wrapper that matches the signature of
    `excel_import_service.parse_daily_excel_report()`. Accepts raw bytes and
    returns the same dict shape that `parse_ddr()` returns."""
    data = parse_ddr(BytesIO(file_bytes))
    # Required by insert_parsed_report()
    if not data["header"].get("well_name"):
        raise ValueError("Could not parse well name from Excel file")
    if not data["header"].get("date"):
        raise ValueError("Could not parse date from Excel file")
    return data


# ---------------------------------------------------------------------------
# JSON CLI
# ---------------------------------------------------------------------------

def _json_safe(obj):
    """Recursively convert datetime / date / time objects to ISO strings."""
    if isinstance(obj, (datetime, date_type)):
        return obj.isoformat()
    if isinstance(obj, time):
        return obj.strftime("%H:%M:%S")
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_json_safe(x) for x in obj]
    return obj


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="Extract structured data from an ENAFOR Daily Drilling Report."
    )
    p.add_argument("source", type=Path, help="Path to the source .xlsx file")
    p.add_argument("-o", "--output", type=Path, default=None,
                   help="Output JSON path (default: <source>.json next to the source)")
    p.add_argument("--pretty", action="store_true",
                   help="Pretty-print the JSON output (indent=2)")
    args = p.parse_args(argv)

    if not args.source.exists():
        sys.exit(f"ERROR: source file not found: {args.source}")

    data = parse_ddr(args.source)

    if args.output is None:
        args.output = args.source.with_suffix(".json")

    with open(args.output, "w", encoding="utf-8") as fh:
        json.dump(_json_safe(data), fh,
                  indent=2 if args.pretty else None,
                  ensure_ascii=False)

    h = data["header"]
    print(f"Wrote {args.output}")
    print(f"  well={h.get('well_name')!r}  rig={h.get('rig_name')!r}  "
          f"field={h.get('field_name')!r}  date={h.get('date')}")
    print(f"  {len(data['activities'])} activities | "
          f"tarif totals: {data['tarif_totals']}")
    print(f"  {len(data['personnel_data'])} personnel rows | "
          f"{len(data['mud_checks'])} mud props | "
          f"safety: {data['safety']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())