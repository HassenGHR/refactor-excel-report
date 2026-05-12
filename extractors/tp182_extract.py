#!/usr/bin/env python3
"""
tp182_extract.py — extract a TP-182 (WIH well, ENTP rig) Daily Workover
Report into the standard dict shape used by ddr_extract.py / tp179_extract.py.

Source layout
-------------
SONATRACH PRODUCTION format, ~110 rows × 17 cols. Header is "WELL : ... |
DATE : ... | TOTAL MD: ... | TOTAL TVD: ... | FORMATION TOP: ... | REP N° :
..." at row 5, with most cells containing both LABEL and value in a single
combined string ("LABEL: value").
Mud panel at rows 11-29 (cols J-L left side, M-N right side).
Mud chemicals at rows 47-55 (cols B-H left side, I-P right side, ~18 items).
Personnel at rows 67-69 (cols B-I left side, J-P right side, ~6 entries
including "WATER TRUCK DRIVER" which counts as water_truck).
Operations table at row 72+ (FROM/TO/HRS/DESCRIPTION/BILL/COMPANY).
Text sections at rows 91-98 (ACTUEL OPERATIONS / PLAN OPERATION / REQUIREMENTS / REMARKS).
"""
from __future__ import annotations
import re
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers (same shape as tp179_extract / ddr_extract)
# ---------------------------------------------------------------------------
def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _float(v, default=0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = _clean(v).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "None"):
        return default
    # Strip trailing units (m, m³, days, Hrs, etc.)
    s = re.sub(r"[a-zA-Zé°%/³]+$", "", s).strip()
    try:
        return float(s)
    except ValueError:
        return default


def _int(v, default=0) -> int:
    return int(_float(v, float(default)))


def _date_parse(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date_type):
        return v
    s = _clean(v)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d,%m,%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _time_parse(v):
    if v is None:
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        # Excel sometimes encodes 00:00 as 1900-01-01 00:00:00 (epoch zero)
        if v.year == 1900 and v.month == 1 and v.day == 1:
            return v.time()
        return v.time()
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        return time((total // 3600) % 24, (total % 3600) // 60)
    s = _clean(v)
    if not s or s in ("-", "None"):
        return None
    if s in ("24:00", "24:00:00"):
        return time(0, 0)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _build_merged_lookup(ws):
    lookup = {}
    for mr in ws.merged_cells.ranges:
        av = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = av
    return lookup


def _cell(ws, r, c, lookup):
    v = ws.cell(r, c).value
    return v if v is not None else lookup.get((r, c))


# In this format the labels and values share a single cell:
#   "WELL : WIH 05"   "DATE : 10-05-2026"   "TOTAL MD: 1592m"   etc.
# We strip the known prefix to get just the value.
def _strip_label(text: str, *labels: str) -> str:
    s = _clean(text)
    for lab in labels:
        rx = rf"^\s*{re.escape(lab)}\s*:?\s*"
        new = re.sub(rx, "", s, flags=re.IGNORECASE)
        if new != s:
            return new.strip()
    return s


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------
def parse_tp182(source: Union[Path, str, BytesIO]) -> dict:
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    ws = wb.active
    L = _build_merged_lookup(ws)

    # =====================================================================
    # HEADER (rows 5-7) — "LABEL: value" format, mostly merged cells
    # =====================================================================
    header = {}

    # Row 5
    header["well_name"]      = _strip_label(_cell(ws, 5, 2,  L), "WELL")
    header["date"]           = _date_parse(_strip_label(_cell(ws, 5, 4,  L), "DATE"))
    header["well_md"]        = _float(_strip_label(_cell(ws, 5, 7,  L), "TOTAL MD"))
    header["tvd"]            = _float(_strip_label(_cell(ws, 5, 9,  L), "TOTAL TVD"))
    header["formation_top"]  = _strip_label(_cell(ws, 5, 11, L), "FORMATION TOP")
    header["day_number"]     = _int(_strip_label(_cell(ws, 5, 14, L), "REP N°", "REP N"))

    # Row 6
    header["supervisor"]          = _strip_label(_cell(ws, 6, 2,  L), "SUPERVISOR")
    header["rig_name"]            = _strip_label(_cell(ws, 6, 5,  L), "RIG NAME")
    header["field_name"]          = _strip_label(_cell(ws, 6, 8,  L), "FIELED", "FIELD")
    header["superintendent"]      = _strip_label(_cell(ws, 6, 10, L), "SUPERINTANDANT",
                                                                       "SUPERINTENDANT",
                                                                       "SUPERINTENDENT")
    raw_acc = _strip_label(_cell(ws, 6, 13, L), "ACC FREE")
    header["accident_free_days"]  = _int(raw_acc)
    header["daily_npt"]           = _float(_strip_label(_cell(ws, 6, 15, L), "Daily NPT"))

    # Row 7
    header["last_formation_test"] = _strip_label(_cell(ws, 7, 2,  L), "LAST FORMATION TEST")
    # The frontend's "Last CSG SHOE" reads from DB column lastCSNlnrSHOE which the
    # insert function fills from header.last_csg_shoe_real OR header.top_shoe.
    # TP-182 has only ONE casing shoe entry (no separate liner), so we put it
    # in `top_shoe` so the frontend "Last CSG SHOE" field fills.
    header["top_shoe"]            = _strip_label(_cell(ws, 7, 5,  L), "LAST CSG SHOE")
    header["top_window"]          = _strip_label(_cell(ws, 7, 8,  L), '9"5/8 TOP WONDOW',
                                                                       "TOP WONDOW",
                                                                       "TOP WINDOW")
    header["bop_test"]            = _date_parse(_strip_label(_cell(ws, 7, 10, L), "LAST BOP TEST"))
    header["last_safety_meeting"] = _date_parse(_strip_label(_cell(ws, 7, 12, L), "LAST SAFETY MEETING"))
    header["cum_npt"]             = _float(_strip_label(_cell(ws, 7, 15, L), "Cum NPT"))

    # =====================================================================
    # LAST BHA (rows 13-29 cols F/H/I) — text summary
    # =====================================================================
    bha_parts = []
    for row in range(13, 30):
        name = _clean(_cell(ws, row, 6, L) or "")        # col F
        if not name or name.upper() == "BHA TOTAL":
            continue
        od     = _clean(_cell(ws, row, 8, L) or "")      # col H
        length = _cell(ws, row, 9, L)                    # col I
        if length is not None:
            try: length = f"{float(length):.2f}m"
            except: length = _clean(length)
        parts = [name]
        if od: parts.append(od)
        if length: parts.append(length)
        bha_parts.append(" / ".join(parts))
    if bha_parts:
        header["bha_details"] = " | ".join(bha_parts)

    # =====================================================================
    # MUD CHECKS  (rows 11-29; left = col J label + col L value;
    #              right = col M label + col N or O value)
    # =====================================================================
    mud_checks = {}
    # Mud type comes from the "OIL BASE MUD (OBM)" or similar at M11
    mud_type_raw = _cell(ws, 11, 13, L) or _cell(ws, 11, 14, L)
    if mud_type_raw:
        mud_checks["mud_type"] = _clean(mud_type_raw)

    # Left column mappings: (row, our_dict_key)
    left_mud = [
        (13, "density"),       # DENSITY (sg)
        (14, "depth"),         # DEPTH (m)
        (15, "fl_tmp"),        # FLOW TEMP (°F)
        (16, "fun_vis"),       # FUN VISCOSITY (Sec)
        (17, "pv"),            # PlASTIC VISCOSITY (cp)
        (18, "yp"),            # YP (lbf/100ft2)
        (19, "gel10sec"),      # GEL 0
        (20, "gel10m"),        # GEL 10
        (21, "apl_fl"),        # API FILRAT
        (22, "hpht_fl"),       # HPHT FILTRAT
        (23, "lgs"),           # LGS (%)
        # (24, "lgs_caco3"),    # LGS CaCo3 (%) — not in our standard schema
        (25, "hgs"),           # HGS (%)
        # (26, "asgs"),
        (27, "solid"),         # CORRECTED SOLID (%)
    ]
    for row, key in left_mud:
        v = _cell(ws, row, 12, L)       # col L
        if v is not None:
            mud_checks[key] = _float(v)

    # Right column mappings: (row, our_dict_key)
    right_mud = [
        (13, "pom"),           # POM
        (14, "sand"),          # SAND (%)
        # (15, "solids"),       # SOLIDS (%) — already in left side as 'solid'
        (16, "oil"),           # OIL (%)
        (17, "h2o"),           # WATER (%)
        # (18, "oil_water"),    # OIL/WATER ratio
        # CL WHOLE MUD (g/l), CL WATER PHASE (g/l) — not in standard schema
        (21, "NaCl"),          # NaCL BY WEIGHT (%)
        (22, "es"),            # ES (Volts)
        # BRINE (g/l) — not in schema
        # (24, "caco3"),       # CaCO3 (g/l)
        # (25, "edta"), (26, "agno3") — not in schema
        (27, "e_line"),        # EXCESS LIME (g/l)
    ]
    for row, key in right_mud:
        # Right-side values can be in col N (14), O (15) or even col P (16)
        v = _cell(ws, row, 14, L) or _cell(ws, row, 15, L) or _cell(ws, row, 16, L)
        if v is not None and not isinstance(v, str):
            mud_checks[key] = _float(v)
        elif v is not None:
            # If it's a string and not just a duplicated label, try parsing
            try:
                mud_checks[key] = float(_clean(v).replace(",", "."))
            except (ValueError, TypeError):
                pass

    # =====================================================================
    # MUD VOLUMES — embedded in mud panel at r28-r29
    # =====================================================================
    mud_volume = {}
    obm_surface = _cell(ws, 28, 12, L)         # L28: OBM Surface Vol value
    if obm_surface is not None:
        mud_volume["pits_volume"] = _float(obm_surface)
    well_vol = _cell(ws, 28, 14, L) or _cell(ws, 28, 15, L)   # right-side: Well Vol
    if well_vol is not None and not isinstance(well_vol, str):
        mud_volume["string_volume"] = _float(well_vol)
    # Compute total
    if mud_volume:
        mud_volume["total_volume"] = sum(
            v for v in (mud_volume.get("pits_volume"), mud_volume.get("string_volume"))
            if v is not None
        )

    # =====================================================================
    # FUEL / WATER / Mud Loss (rows 30-32) — extra context, not in schema
    # =====================================================================
    fuel = _cell(ws, 31, 3, L)
    water = _cell(ws, 32, 3, L)
    if fuel is not None:
        header["fuel_vol"] = _float(fuel)
    if water is not None:
        header["water_vol"] = _float(water)

    # =====================================================================
    # MUD CHEMICAL USAGE (rows 47-55)
    # Left side  : col B item, col E units, col F rec'd, col G used, col H end
    # Right side : col I item, col L units, col M rec'd, col O used, col P end
    # =====================================================================
    chemicals = []
    for row in range(47, 56):
        # LEFT block
        item_left = _clean(_cell(ws, row, 2, L) or "")
        if item_left:
            rec   = _cell(ws, row, 6, L)
            used  = _cell(ws, row, 7, L)
            end   = _cell(ws, row, 8, L)
            if any(v is not None for v in (rec, used, end)):
                chemicals.append({
                    "item":     item_left,
                    "units":    _clean(_cell(ws, row, 5, L) or ""),
                    "received": _clean(rec  if rec  is not None else ""),
                    "used":     _clean(used if used is not None else ""),
                    "on_loc":   _clean(end  if end  is not None else ""),
                })
            else:
                # Include the chemical even with no usage info — it's on rig
                chemicals.append({
                    "item":   item_left,
                    "units":  _clean(_cell(ws, row, 5, L) or ""),
                    "received":"", "used":"", "on_loc":"0",
                })
        # RIGHT block
        item_right = _clean(_cell(ws, row, 9, L) or "")
        if item_right:
            rec   = _cell(ws, row, 13, L)
            used  = _cell(ws, row, 15, L)
            end   = _cell(ws, row, 16, L)
            if any(v is not None for v in (rec, used, end)):
                chemicals.append({
                    "item":     item_right,
                    "units":    _clean(_cell(ws, row, 12, L) or ""),
                    "received": _clean(rec  if rec  is not None else ""),
                    "used":     _clean(used if used is not None else ""),
                    "on_loc":   _clean(end  if end  is not None else ""),
                })
            else:
                chemicals.append({
                    "item":   item_right,
                    "units":  _clean(_cell(ws, row, 12, L) or ""),
                    "received":"", "used":"", "on_loc":"0",
                })

    # =====================================================================
    # PERSONNEL  (rows 67-69)
    # Left  : col B company, col D number, col G names
    # Right : col J company, col L number, col N names
    # The "WATER TRUCK DRIVER" row also gives us the water_truck count.
    # =====================================================================
    personnel = []
    for row in range(67, 71):
        # LEFT block
        company_l = _clean(_cell(ws, row, 2, L) or "")
        if company_l:
            count_l = _cell(ws, row, 4, L)
            names_l = _clean(_cell(ws, row, 7, L) or "")
            if "WATER TRUCK" in company_l.upper() and count_l is not None:
                header["water_truck"] = _int(count_l)
            personnel.append({
                "company": company_l,
                "number":  _int(count_l, 0),
                "hours":   "",
                "names":   names_l,
            })
        # RIGHT block
        company_r = _clean(_cell(ws, row, 10, L) or "")
        if company_r:
            count_r = _cell(ws, row, 12, L)
            names_r = _clean(_cell(ws, row, 14, L) or "")
            personnel.append({
                "company": company_r,
                "number":  _int(count_r, 0),
                "hours":   "",
                "names":   names_r,
            })

    # =====================================================================
    # OPERATIONS  (rows 73+)
    # Header at r72: FROM(B) | TO(C) | HRS(D) | DESCRIPTION(E-N) | BILL(O) | COMPANY(P)
    # We treat a row as a real operation only if it has BILL or HRS set.
    # =====================================================================
    activities = []
    for row in range(73, 90):
        start = _cell(ws, row, 2, L)
        end   = _cell(ws, row, 3, L)
        hrs   = _cell(ws, row, 4, L)
        desc  = _clean(_cell(ws, row, 5, L) or "")
        bill  = _clean(_cell(ws, row, 15, L) or "")
        comp  = _clean(_cell(ws, row, 16, L) or "")

        # Skip empty rows
        if not bill and not desc:
            continue
        # Real operation rows have a BILL code (T1/T2/T3/FT/NR).  Rows that
        # have a time range but no BILL are display-only / overflow text and
        # are already counted within an earlier row's 24h block.
        if not bill:
            continue
        # Skip the summary line (contains "T1=" or "T2=" markers)
        if any(t in desc.upper() for t in ("T1=", "T2=", "T3=", "FT=", "NR=")):
            continue

        start_t = _time_parse(start)
        end_t   = _time_parse(end)
        hours = 0.0
        if isinstance(hrs, (int, float)):
            hours = float(hrs)
        elif isinstance(hrs, (datetime, time, timedelta)):
            ht = _time_parse(hrs) if not isinstance(hrs, timedelta) else None
            if ht is not None:
                hours = ht.hour + ht.minute / 60.0
                # Excel epoch zero (1900-01-01 00:00) means a 24-hour duration
                if hours == 0.0 and start_t and end_t and start_t == end_t == time(0, 0):
                    hours = 24.0
        elif start_t and end_t:
            sm = start_t.hour * 60 + start_t.minute
            em = end_t.hour * 60 + end_t.minute
            if em == sm: hours = 24.0
            elif em > sm: hours = (em - sm) / 60.0
            else: hours = (em + 1440 - sm) / 60.0

        activities.append({
            "start_time": start_t,
            "end_time":   end_t,
            "hours":      hours,
            "phase_name": "",
            "code":       "",
            "sub":        "",
            "description": desc,
            "start_md": 0, "end_md": 0,
            "npt": 0, "npt_detail": "",
            "npt_company": "", "op_company": comp,
            "bill": bill,
        })

    # =====================================================================
    # TEXT SECTIONS  (rows 91-98)
    # =====================================================================
    text_sections = {}
    for row in range(91, 100):
        v = _cell(ws, row, 2, L)
        if v is None: continue
        text = _clean(v)
        if not text: continue
        upper = text.upper()
        if upper.startswith("ACTUEL OPERATIONS") or "ACTUEL OPERATIONS" in upper[:30]:
            value = re.sub(r"^ACTUEL\s+OPERATIONS\s*\d*[Hh]?\d*\s*:?\s*", "", text, flags=re.IGNORECASE)
            text_sections["current_operation"] = value
            text_sections["day_summary"]       = value
        elif upper.startswith("PLAN OPERATION"):
            text_sections["plan_operations"] = re.sub(r"^PLAN\s+OPERATION\s*:?\s*", "", text, flags=re.IGNORECASE)
        elif "REQUIREMENT" in upper or "RENTAL EQUIPMENT" in upper:
            v2 = re.sub(r"^REQUIREMENTS\s*/\s*RENTAL EQUIPMENT\s*:?\s*", "", text, flags=re.IGNORECASE)
            if v2: text_sections["requirements"] = v2
        elif upper.startswith("REMARK"):
            v2 = re.sub(r"^REMARKS?\s*:?\s*", "", text, flags=re.IGNORECASE)
            if v2: text_sections["remarks"] = v2

    # =====================================================================
    # SAFETY
    # =====================================================================
    safety = {}
    if header.get("accident_free_days") is not None:
        safety["accident_free_days"] = header["accident_free_days"]
    if header.get("water_truck") is not None:
        safety["water_truck"] = header["water_truck"]
    if header.get("last_safety_meeting"):
        safety["last_safety_meeting"] = header["last_safety_meeting"]

    wb.close()

    return {
        "header": header,
        "activities": activities,
        "text_sections": text_sections,
        "mud_checks": mud_checks,
        "mud_volume": mud_volume,
        "mud_chemical_usage": chemicals,
        "personnel_data": personnel,
        "pumps": [],
        "well_location": {},
        "survey_data": [],
        "safety": safety,
        "tarif_totals": {},
    }


# Drop-in compat for the converter's import
parse_daily_excel_report = parse_tp182


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Usage: tp182_extract.py SOURCE.xlsx")
    data = parse_tp182(Path(sys.argv[1]))

    def default(o):
        if isinstance(o, (date_type, datetime)): return o.isoformat()
        if isinstance(o, time): return o.strftime("%H:%M:%S")
        if isinstance(o, timedelta): return o.total_seconds()
        return str(o)
    print(json.dumps(data, indent=2, default=default, ensure_ascii=False))