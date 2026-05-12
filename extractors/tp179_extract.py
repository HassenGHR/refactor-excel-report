#!/usr/bin/env python3
"""
tp179_extract.py — extract a TP-179 (HTJW well, ENTP rig) Daily Workover
Report into the same dict shape produced by ddr_extract.py.

Designed to be a drop-in alternative: anywhere ddr_extract.parse_ddr() is
imported, you can use tp179_extract.parse_tp179() with the same downstream
pipeline (to_router_excel → existing TP parser → patched insert_parsed_report).

Source layout
-------------
Single sheet, ~44 rows × 28 cols, French/English mix, "RAPPORT JOURNALIER
DU WORK OVER" at H2. Header is in rows 2-7, operations table at rows 12-29,
mud panel at rows 7-13 cols S-T, mud volumes at rows 19-21, chemicals at
rows 24-34, personnel at rows 28-33, and Représentant maître d'œuvre /
camions citernes at rows 35-36.
"""
from __future__ import annotations
import re
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Small helpers (same shape as ddr_extract.py so the rest of the pipeline
# doesn't see any difference)
# ---------------------------------------------------------------------------
def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _float(v, default=0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = _clean(v).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "None"):
        return default
    s = re.sub(r"[a-zA-Z%]+$", "", s).strip()
    try:
        return float(s)
    except ValueError:
        return default


def _int(v, default=0) -> int:
    return int(_float(v, float(default)))


def _date_parse(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date_type):
        return v
    s = _clean(v)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d,%m,%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _time_parse(v):
    if v is None:
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        return time((total // 3600) % 24, (total % 3600) // 60)
    s = _clean(v)
    if not s or s in ("-", "None"):
        return None
    if s in ("24:00", "24:00:00"):
        return time(0, 0)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _build_merged_lookup(ws):
    lookup = {}
    for mr in ws.merged_cells.ranges:
        anchor_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = anchor_val
    return lookup


def _cell(ws, r, c, lookup):
    v = ws.cell(r, c).value
    return v if v is not None else lookup.get((r, c))


def _strip_prefix(text: str, *prefixes: str) -> str:
    """Strip any of the given case-insensitive prefixes (with optional colon)."""
    s = _clean(text)
    for p in prefixes:
        # Try exact prefix, with or without colon and whitespace
        rx = rf"^\s*{re.escape(p)}\s*:?\s*"
        new = re.sub(rx, "", s, flags=re.IGNORECASE)
        if new != s:
            return new.strip()
    return s


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------
def parse_tp179(source: Union[Path, str, BytesIO]) -> dict:
    """Extract a TP-179 daily workover report into the standard dict shape."""
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    ws = wb.active
    L = _build_merged_lookup(ws)

    # =====================================================================
    # HEADER
    # =====================================================================
    header = {}

    # Well name from A6 with "Well :" prefix
    well_raw = _cell(ws, 6, 1, L)        # A6
    header["well_name"] = _strip_prefix(_clean(well_raw), "Well", "Puits")

    # Rig name from F6 with "RIG :" prefix
    rig_raw = _cell(ws, 6, 6, L)         # F6
    header["rig_name"] = _strip_prefix(_clean(rig_raw), "RIG", "APPAREIL")

    # Field / Zone — J6
    header["field_name"] = _clean(_cell(ws, 6, 10, L))

    # Class (extra context) — H6
    header["well_class"] = _clean(_cell(ws, 6, 8, L))

    # Report date — S4
    header["date"] = _date_parse(_cell(ws, 4, 19, L))

    # Day number — U2
    header["day_number"] = _int(_cell(ws, 2, 21, L))

    # BOP tests — K7 (last), N7 (next)
    header["bop_test"] = _date_parse(_cell(ws, 7, 11, L))
    header["next_bop_test"] = _date_parse(_cell(ws, 7, 14, L))

    # Workover objective — A10 (label is A8 "BUT DU WORK OVER")
    header["well_objective"] = _clean(_cell(ws, 10, 1, L))

    # Mud type — strip "TYPE:" prefix from S6
    mud_type_raw = _cell(ws, 6, 19, L)
    mud_type = _strip_prefix(_clean(mud_type_raw), "TYPE")

    # Supervisor — S36 (value under "Représentant maître d'œuvre:" at S35).
    # Per the user's convention, this is a supervisor (12-hour shift), not a
    # superintendent.  The source only has one name on this rig-move day, so
    # the `superintendent` slot stays empty.
    sup_label = _clean(_cell(ws, 35, 19, L) or "")
    if "REPRÉSENTANT" in sup_label.upper() or "MAITRE" in sup_label.upper() or "MAÎTRE" in sup_label.upper():
        sup_val = _cell(ws, 36, 19, L)
        if sup_val:
            header["supervisor"] = _clean(sup_val)

    # Water truck count — R36 (the "CAMIONS CITERNE" row)
    if "CAMIONS" in _clean(_cell(ws, 36, 11, L) or "").upper():
        header["water_truck"] = _int(_cell(ws, 36, 18, L))

    # Casing info is NOT present in this layout — leave unset.

    # =====================================================================
    # OPERATIONS — rows 13+ cols A/B/C/D, no per-row hours column
    # Tarif daily totals at row 10 cols L-P
    # =====================================================================
    activities = []
    daily_totals = {}
    for code, col in (("T1", 12), ("T2", 13), ("T3", 14), ("T4", 15), ("NR", 16)):
        v = _float(_cell(ws, 10, col, L), 0)
        if v:
            daily_totals[code] = v

    # Walk operation rows. The layout uses ONE row per operation with
    # A=start time, B=end time, C=tarif, D=description. Hours come from
    # the daily-totals table since there's no per-row H column.
    for row in range(13, 30):
        a = _cell(ws, row, 1, L)
        b = _cell(ws, row, 2, L)
        c = _cell(ws, row, 3, L)
        d = _cell(ws, row, 4, L)
        # Skip rows that are labels or blank
        if a is None and not d:
            continue
        # Some rows in this region are non-operations (e.g., "BUT DU WORK
        # OVER" overflow, "Situation" rows starting at r30). Filter by tarif.
        tarif = _clean(c).upper()
        if not re.match(r"^T[1-4]$|^NR$", tarif):
            continue

        start_t = _time_parse(a)
        end_t = _time_parse(b)
        if start_t is None and end_t is None and not d:
            continue

        # Compute duration. 00:00 → 00:00 means a 24h operation.
        if start_t and end_t:
            sm = start_t.hour * 60 + start_t.minute
            em = end_t.hour * 60 + end_t.minute
            if em == sm:
                hours = 24.0
            elif em > sm:
                hours = (em - sm) / 60.0
            else:
                hours = (em + 24 * 60 - sm) / 60.0
        else:
            # Fall back to the tarif's daily total when only one operation has
            # that tarif code today (common in workover rig-move days).
            same_code = [op for op in activities if op.get("bill") == tarif]
            if not same_code and tarif in daily_totals:
                hours = daily_totals[tarif]
            else:
                hours = 0.0

        activities.append({
            "start_time": start_t,
            "end_time": end_t,
            "hours": hours,
            "phase_name": "",
            "code": "",
            "sub": "",
            "description": _clean(d),
            "start_md": 0, "end_md": 0,
            "npt": 0, "npt_detail": "",
            "npt_company": "", "op_company": "",
            "bill": tarif,
        })

    # =====================================================================
    # TEXT SECTIONS (Situation, Programme prévu, Remarque)
    # =====================================================================
    text_sections = {}
    # Scan column A rows 28-40 for "SITUATION", "PROGRAMME", "REMARQUE"
    for row in range(28, 41):
        label = _clean(_cell(ws, row, 1, L) or "").upper()
        if not label:
            continue
        # Value lives in column E (col 5) on this layout (merged across to J)
        val = _clean(_cell(ws, row, 5, L) or "")
        if "SITUATION" in label:
            text_sections["current_operation"] = val
            text_sections["day_summary"] = val
        elif "PROGRAMME" in label or "PLAN" in label:
            text_sections["plan_operations"] = val
        elif "REMARQUE" in label or label.startswith("REMARK"):
            if val:
                text_sections["remarks"] = val

    # =====================================================================
    # MUD CHECKS  (rows 7-13 col S label, col T value)
    # =====================================================================
    mud_checks = {}
    if mud_type:
        mud_checks["mud_type"] = mud_type
    mud_map = [
        (7,  "density"),       # Densité (sg)
        (8,  "fun_vis"),       # Funnel Visco
        (9,  "pv"),            # VP (cp)
        (10, "yp"),            # YP (lb/100ft)
        (11, "apl_fl"),        # Filtrat
        (12, "gel10sec"),      # Gel 10
        (13, "gel10m"),        # Gel 30 (closest match in our dict shape)
    ]
    for row, key in mud_map:
        v = _cell(ws, row, 20, L)        # col T
        if v is None:
            continue
        # Keep 0 values too — they are meaningful ("we measured and it was 0").
        mud_checks[key] = _float(v)

    # =====================================================================
    # MUD VOLUMES  (rows 19-21 col S label, col T value)
    # =====================================================================
    mud_volume = {}
    for row, key in [(19, "string_volume"),     # V. puits
                     (20, "pits_volume"),       # V. surface
                     (21, "reserve_volume")]:   # V. Reserve
        v = _cell(ws, row, 20, L)
        if v is not None:
            mud_volume[key] = _float(v)
    # The reports doesn't show an explicit "total" — we sum.
    if mud_volume:
        mud_volume.setdefault("total_volume",
                              sum(v for v in (mud_volume.get("string_volume"),
                                              mud_volume.get("pits_volume"),
                                              mud_volume.get("reserve_volume"))
                                  if v))

    # =====================================================================
    # MUD CHEMICAL USAGE  (rows 24-34 cols S/T/U/V)
    # =====================================================================
    chemicals = []
    for row in range(24, 35):
        item = _clean(_cell(ws, row, 19, L) or "")
        if not item or "STOCK" in item.upper():
            continue
        initial = _cell(ws, row, 20, L)
        used = _cell(ws, row, 21, L)
        final = _cell(ws, row, 22, L)
        # Skip ONLY when every cell is truly empty (None).  A row with
        # explicit 0 values means "this chemical is on the rig, current
        # stock is zero" — still worth recording.
        if all(v is None for v in (initial, used, final)):
            continue
        chemicals.append({
            "item": item,
            "units": "",
            "received": _clean(initial if initial is not None else ""),
            "used":     _clean(used    if used    is not None else ""),
            "on_loc":   _clean(final   if final   is not None else ""),
        })

    # =====================================================================
    # PERSONNEL  (rows 28-33 cols K=role, N=company, R=count)
    # =====================================================================
    personnel = []
    for row in range(28, 34):
        role = _clean(_cell(ws, row, 11, L) or "")
        if not role or "TOTAL" in role.upper():
            continue
        company = _clean(_cell(ws, row, 14, L) or "")
        count = _cell(ws, row, 18, L)
        if not company and count in (None, "", 0):
            continue
        personnel.append({
            "company": role,            # "Maitre d'œuvre", "Chauffeurs..." etc.
            "number": _int(count, 0),
            "hours": "",
            "names": company,
        })

    # =====================================================================
    # SAFETY counters — this layout doesn't carry an accident-free counter
    # =====================================================================
    safety = {}
    if header.get("water_truck") is not None:
        safety["water_truck"] = header["water_truck"]

    # =====================================================================
    # TARIF totals (extra info)
    # =====================================================================
    tarif_totals = dict(daily_totals)

    wb.close()

    return {
        "header": header,
        "activities": activities,
        "text_sections": text_sections,
        "mud_checks": mud_checks,
        "mud_volume": mud_volume,
        "mud_chemical_usage": chemicals,
        "personnel_data": personnel,
        "pumps": [],
        "well_location": {},
        "survey_data": [],
        "safety": safety,
        "tarif_totals": tarif_totals,
    }


# Drop-in compat for the converter's import
parse_daily_excel_report = parse_tp179
parse_ddr = parse_tp179


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Usage: tp179_extract.py SOURCE.xlsx")
    data = parse_tp179(Path(sys.argv[1]))
    # Pretty-print the dict shape
    def default(o):
        if isinstance(o, (date_type, datetime)): return o.isoformat()
        if isinstance(o, time): return o.strftime("%H:%M:%S")
        if isinstance(o, timedelta): return o.total_seconds()
        return str(o)
    print(json.dumps(data, indent=2, default=default))