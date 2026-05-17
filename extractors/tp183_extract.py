#!/usr/bin/env python3
"""
tp183_extract.py — extract a TP-183 (ENTP rig 183, TMLS wells) Daily
Workover Report into the standard dict shape.

Source layout
-------------
Compact single-sheet French template (~62 rows × 19 cols).  Title at D1:
"RAPPORT JOURNALIER WORK OVER" (no "DE", different from TP-179/TP-173/ENF#04).

Header (rows 3-4):
  A3=PUITS    A4=well name (TMLS 6)
  C3=APPAREIL C4=rig name  (TP 183)
  D3=Dernier Tubage  D4=CSG description (7" P110 N.VAM...)
                     E4=CSG depth (2790)
  H3=Liner 4"1/2 (label only on this template)
  J3=Fond label   J4=TD value (3080)
  K3=TYPE BOUE    M3=mud type (OBM)

Tarif totals at row 7 (F=T1, H=T2, I=T3, J=T4 — note gap at col G):
  F7=21  H7=3  I7=0  J7=  (E7=H/Jour label)
Daily cost at E11; cumul cost in nearby H-K columns.

Mud panel at K9-L10 (density, viscosity).
Volumes at N7/N8 (Vol puits / Vol surf).

Activities at rows 10-12: A=start, B=end, C=description.  Some time values
are str ("00:00", "24h00") rather than datetime.time — need to handle.

Cost breakdown E13-26: E=company label, I=montant DA.

Chemicals K17-32: K=item, L=units, M=used, N=stock.

Personnel E28-34: E=role label, J=count.

Situation A37, Programme A38.

Supervisors K37 + M37 — TWO names (Rep maître œuvre + Resp sce puits) →
both go to supervisor + superintendent per the two-supervisor convention.

Distinguishing marker (from other "RAPPORT JOURNALIER" templates):
    - title "RAPPORT JOURNALIER WORK OVER" (no "DE")
    - "TP 183" or "TP-183" anywhere in the markers
    - well "TMLS" prefix
"""
from __future__ import annotations
import re
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Union, List

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers (same shape as the other extractors)
# ---------------------------------------------------------------------------
def _clean(v) -> str:
    if v is None: return ""
    s = str(v)
    # Strip Excel error tokens — these come from broken cross-sheet formulas
    if s.strip() in ("#VALUE!", "#REF!", "#NAME?", "#N/A", "#DIV/0!", "#NULL!"):
        return ""
    return re.sub(r"\s+", " ", s).strip()


def _float(v, default=0.0) -> float:
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    s = _clean(v).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "None"): return default
    s = re.sub(r"[a-zA-Zé°%/³]+$", "", s).strip()
    try: return float(s)
    except ValueError: return default


def _int(v, default=0) -> int:
    return int(_float(v, float(default)))


def _date_parse(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date_type): return v
    s = _clean(v)
    if not s or s in ("-", "/", "--", "//"): return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None


def _time_parse(v):
    """Parse a time value.  Handles datetime.time, str '24h00' / '00:00',
    and Excel-epoch datetime (1900-01-01 means 24:00)."""
    if v is None: return None
    if isinstance(v, time): return v
    if isinstance(v, datetime):
        if v.year == 1900 and v.month == 1 and v.day == 1:
            return v.time()
        return v.time()
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        if total == 86400: return time(0, 0)
        return time((total // 3600) % 24, (total % 3600) // 60)
    s = _clean(v)
    if not s or s in ("-", "None"): return None
    if s in ("24:00", "24:00:00", "24h00", "24H00", "24h"): return time(0, 0)
    # Handle "13h30" / "13H30" style
    m = re.match(r"^(\d{1,2})[hH](\d{0,2})$", s)
    if m:
        return time(int(m.group(1)) % 24, int(m.group(2) or 0))
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try: return datetime.strptime(s, fmt).time()
        except ValueError: continue
    return None


def _duration_hours(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, timedelta): return v.total_seconds() / 3600.0
    if isinstance(v, time):     return v.hour + v.minute / 60.0
    if isinstance(v, datetime):
        if v.year == 1900 and v.month == 1 and v.day == 1:
            return v.hour + v.minute / 60.0
        return 0.0
    return _float(v)


def _build_merged_lookup(ws):
    lookup = {}
    for mr in ws.merged_cells.ranges:
        av = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = av
    return lookup


def _cell(ws, r, c, lookup):
    v = ws.cell(r, c).value
    return v if v is not None else lookup.get((r, c))


def _strip_label(text: str, *labels: str) -> str:
    """Strip a known label prefix from a string.  Used for cells like
    'BUT  DU  WORK  OVER : \\n Reprise technique...' where label+value share
    a cell."""
    s = _clean(text)
    for lab in labels:
        rx = rf"^\s*{re.escape(lab)}\s*:?\s*"
        new = re.sub(rx, "", s, flags=re.IGNORECASE)
        if new != s:
            return new.strip()
    return s


def _normalize_rig(name: str) -> str:
    """'TP 183' / 'TP-183' / 'TP183' → 'TP-183' (canonical form)."""
    s = _clean(name).upper()
    m = re.search(r"TP[-\s]*(\d+)", s)
    if m: return f"TP-{m.group(1)}"
    return _clean(name)


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------
def parse_tp183(source: Union[Path, str, BytesIO]) -> dict:
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    ws = wb.active
    L = _build_merged_lookup(ws)

    # =====================================================================
    # HEADER (rows 3-4)
    # =====================================================================
    header = {}

    # Well  — A4
    header["well_name"]  = _clean(_cell(ws, 4, 1, L))
    # Rig   — C4
    header["rig_name"]   = _normalize_rig(_cell(ws, 4, 3, L) or "")
    # CSG (Dernier Tubage) — D4 description + E4 depth
    csg_desc = _clean(_cell(ws, 4, 4, L) or "")
    csg_depth = _cell(ws, 4, 5, L)
    if csg_desc or csg_depth is not None:
        if csg_depth is not None:
            try:
                d = int(float(str(csg_depth)))
                header["top_shoe"] = f"{csg_desc} @ {d}m" if csg_desc else f"@ {d}m"
            except (ValueError, TypeError):
                header["top_shoe"] = f"{csg_desc} @ {_clean(csg_depth)}".strip()
        else:
            header["top_shoe"] = csg_desc

    # Liner top — H3 ("Liner 4\"1/2" label).  This template doesn't always
    # carry the liner depth; if present, it's in J4 next to the FOND label.
    # We keep just the size string when no depth is given.
    liner = _clean(_cell(ws, 3, 8, L) or "")
    if liner and "LINER" in liner.upper():
        size = re.sub(r"^\s*LINER\s+", "", liner, flags=re.IGNORECASE).strip()
        if size:
            header["last_csg_top"] = size

    # Total Depth (FOND) — J4
    fond = _cell(ws, 4, 10, L)
    if fond is not None:
        header["well_md"] = _float(fond)

    # Mud type — M3
    mud_type = _clean(_cell(ws, 3, 13, L) or "")

    # Date — K2 (label "JOURNEE DU" sits at K1).  In live deployments
    # this cell holds a resolved date like 2026-05-10.  In standalone
    # uploads where the cross-sheet formula breaks it resolves to
    # "#VALUE!" — _clean / _date_parse both treat that as None, so we
    # gracefully skip and let to_router_excel.py fall back to the
    # filename / --date override.
    date_raw = _cell(ws, 2, 11, L)                # K2
    d = _date_parse(date_raw)
    if d:
        header["date"] = d

    # Day number — read from the SHEET NAME ("RAP N°46").  Earlier I
    # tried K2 for this, but K2 actually holds the date, not the day
    # number — the day number lives in the sheet title.
    sheet_match = re.search(r"N\s*[°o]\s*(\d+)", ws.title or "", re.IGNORECASE)
    if sheet_match:
        header["day_number"] = int(sheet_match.group(1))

    # Objective — A5/A6 (label+value share the cell)
    obj_raw = _cell(ws, 6, 1, L) or _cell(ws, 5, 1, L)
    if obj_raw:
        header["well_objective"] = _strip_label(str(obj_raw), "BUT DU WORK OVER",
                                                "BUT  DU  WORK  OVER")

    # =====================================================================
    # TARIF TOTALS (row 7: F=T1, H=T2, I=T3, J=T4 — note gap at col G)
    # Row 8: H/Cum gives cumulative — we capture both
    # =====================================================================
    tarif_totals = {}
    for code, col in [("t1", 6), ("t2", 8), ("t3", 9), ("t4", 10)]:
        v = _cell(ws, 7, col, L)
        if v is not None:
            try:
                hrs = _float(v)
                if hrs > 0:
                    tarif_totals[code] = hrs
            except Exception:
                pass

    # =====================================================================
    # COSTS — daily at E11 (after the JOURNALIER label at E10),
    # cumul cost in H11 (after CUMULE label at H10)
    # =====================================================================
    daily_cost = _cell(ws, 11, 5, L)
    if daily_cost is not None:
        try:
            v = _float(daily_cost)
            if v > 0: header["daily_cost"] = int(v)
        except Exception:
            pass
    cum_cost = _cell(ws, 11, 8, L)
    if cum_cost is not None:
        try:
            v = _float(cum_cost)
            if v > 0: header["cum_cost"] = int(v)
        except Exception:
            pass

    # =====================================================================
    # OPERATIONS  (rows 10-12 by default; scan a few more rows to be safe.
    # Stop at any row that contains a non-time string in col A and is
    # clearly a section header.)
    # =====================================================================
    activities = []
    OPS_START = 10
    OPS_END = 30      # generous upper bound
    for row in range(OPS_START, OPS_END):
        start = _cell(ws, row, 1, L)
        end   = _cell(ws, row, 2, L)
        desc  = _clean(_cell(ws, row, 3, L) or "")

        # Stop markers — these appear in col A when the ops block ends
        a_text = _clean(_cell(ws, row, 1, L) or "").upper()
        if any(m in a_text for m in (
                "SITUATION", "PROGRAMME", "BUT DU WORK",
                "ANALYSE DES TEMPS", "TOTAL",
        )) and not _time_parse(start):
            break

        # Skip empty rows
        if start is None and end is None and not desc:
            continue

        start_t = _time_parse(start)
        end_t   = _time_parse(end)

        if start_t is not None:
            # Compute hours from times — these are the source of truth
            hours = 0.0
            if start_t and end_t:
                sm = start_t.hour * 60 + start_t.minute
                em = end_t.hour * 60 + end_t.minute
                if em == sm:    hours = 24.0
                elif em > sm:   hours = (em - sm) / 60.0
                else:           hours = (em + 1440 - sm) / 60.0

            activities.append({
                "start_time": start_t,
                "end_time":   end_t,
                "hours":      hours,
                "phase_name": "",
                "code": "", "sub": "",
                "description": desc,
                "start_md": 0, "end_md": 0,
                "npt": 0, "npt_detail": "",
                "npt_company": "", "op_company": "",
                "bill": "",        # this template has no per-op bill code
            })
        elif desc and activities:
            # Continuation row — fold into previous op's description
            prev = activities[-1]
            prev["description"] = (prev["description"] + "\n" + desc).strip()

    # Back-assign bill codes from the tarif daily totals via the shared helper
    from helpers.bill_code_assign import assign_bill_codes
    activities = assign_bill_codes(activities, tarif_totals)

    # =====================================================================
    # MUD CHECKS (K9-L10, plus a few more rows)
    # =====================================================================
    mud_checks = {}
    if mud_type and mud_type.upper() != "TYPE BOUE":
        mud_checks["mud_type"] = mud_type

    # K9=Densité label, L9=value
    dens = _cell(ws, 9, 12, L)
    if dens is not None: mud_checks["density"] = _float(dens)
    # K10=Visc. March., L10
    fv = _cell(ws, 10, 12, L)
    if fv is not None: mud_checks["fun_vis"] = _float(fv)
    # Yield at M9
    yield_v = _cell(ws, 9, 13, L)
    if yield_v is not None:
        try:
            v = _float(yield_v)
            if v > 0: mud_checks["yp"] = v
        except: pass
    # PH at M10
    ph = _cell(ws, 10, 13, L)
    if ph is not None:
        try:
            v = _float(ph)
            if v > 0: mud_checks["ph"] = v
        except: pass

    # =====================================================================
    # MUD VOLUMES — Vol puits=N7, Vol surf=N8 (labels at M7, M8)
    # =====================================================================
    mud_volume = {}
    vp = _cell(ws, 7, 14, L)
    if vp is not None:
        try: mud_volume["string_volume"] = _float(vp)
        except: pass
    vs = _cell(ws, 8, 14, L)
    if vs is not None:
        try: mud_volume["pits_volume"] = _float(vs)
        except: pass
    if mud_volume:
        mud_volume["total_volume"] = sum(
            v for v in (mud_volume.get("string_volume"), mud_volume.get("pits_volume"))
            if isinstance(v, (int, float))
        )

    # =====================================================================
    # MUD CHEMICAL USAGE (rows 18-32; K=item, L=units, M=used, N=stock)
    # =====================================================================
    chemicals = []
    for row in range(18, 36):
        item = _clean(_cell(ws, row, 11, L) or "")
        if not item or item.upper() in ("PRODUITS", "PRODUITS SH DP OHT"):
            continue
        units = _clean(_cell(ws, row, 12, L) or "")
        used  = _cell(ws, row, 13, L)
        stock = _cell(ws, row, 14, L)
        chemicals.append({
            "item":     item,
            "units":    units,
            "received": "",
            "used":     _clean(str(used)  if used  is not None else ""),
            "on_loc":   _clean(str(stock) if stock is not None else ""),
        })

    # =====================================================================
    # PERSONNEL (rows 28-34; col E label, col J count)
    # The TOTAL row at the bottom is excluded.
    # =====================================================================
    personnel = []
    for row in range(28, 36):
        role = _clean(_cell(ws, row, 5, L) or "")
        count = _cell(ws, row, 10, L)
        if not role: continue
        if role.upper().startswith("TOTAL"): continue
        if "PERSONNEL" in role.upper(): continue
        personnel.append({
            "company": role,
            "number":  _int(count, 0),
            "hours":   "",
            "names":   "",
        })

    # =====================================================================
    # SUPERVISORS  (row 36-37 in cols K, M)
    # Layout: K36 "Rep maître œuvre" label, M36 "Resp sce puits" label;
    # K37 supervisor name, M37 second supervisor name.
    # Per the two-supervisor convention, BOTH names land in the
    # supervisor / superintendent slots.
    # =====================================================================
    sup1_label = _clean(_cell(ws, 36, 11, L) or "")
    sup2_label = _clean(_cell(ws, 36, 13, L) or "")
    sup1_name  = _clean(_cell(ws, 37, 11, L) or "")
    sup2_name  = _clean(_cell(ws, 37, 13, L) or "")
    if sup1_name and "MAÎTRE" in sup1_label.upper() or "MAITRE" in sup1_label.upper() or sup1_name:
        # First supervisor → "supervisor"
        if sup1_name and sup1_name.upper() not in ("REP MAÎTRE ŒUVRE", "REP MAITRE OEUVRE",
                                                     "MAÎTRE D'ŒUVRE"):
            header["supervisor"] = sup1_name
    if sup2_name and sup2_name.upper() not in ("RESP SCE PUITS",):
        # Second supervisor → "superintendent" slot (column-name misnomer,
        # but per the convention these are both 12-hour shift supervisors).
        header["superintendent"] = sup2_name

    # =====================================================================
    # TEXT SECTIONS (rows 37-40)
    # SITUATION AU RAPPORT: ... at A37
    # PROGRAMME PRÉVU: ...   at A38
    # =====================================================================
    text_sections = {}
    sit = _cell(ws, 37, 1, L)
    if sit:
        v = _strip_label(str(sit), "SITUATION AU RAPPORT", "SITUATION")
        if v:
            # Cap at 300 chars to fit the frontend display limit
            if len(v) > 300:
                v = v[:300].rsplit(None, 1)[0] + "…"
            text_sections["current_operation"] = v
            text_sections["day_summary"]       = v

    plan = _cell(ws, 38, 1, L)
    if plan:
        v = _strip_label(str(plan), "PROGRAMME PRÉVU", "PROGRAMME PREVU", "PROGRAMME")
        if v:
            text_sections["plan_operations"] = v

    wb.close()

    # Track cost breakdown by company in a side dict (not used by current
    # downstream, but available if you ever wire up per-vendor cost reporting)
    cost_breakdown = {}
    for row in range(13, 27):
        co = _clean(_cell(ws, row, 5, L) or "") if False else None
    # (skipped — we already closed wb above; leave hook for future expansion)

    return {
        "header": header,
        "activities": activities,
        "text_sections": text_sections,
        "mud_checks": mud_checks,
        "mud_volume": mud_volume,
        "mud_chemical_usage": chemicals,
        "personnel_data": personnel,
        "pumps": [],
        "well_location": {},
        "survey_data": [],
        "safety": {},
        "tarif_totals": tarif_totals,
    }


# Drop-in compat
parse_daily_excel_report = parse_tp183


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Usage: tp183_extract.py SOURCE.xlsx")
    data = parse_tp183(Path(sys.argv[1]))

    def default(o):
        if isinstance(o, (date_type, datetime)): return o.isoformat()
        if isinstance(o, time): return o.strftime("%H:%M:%S")
        if isinstance(o, timedelta): return o.total_seconds()
        return str(o)
    print(json.dumps(data, indent=2, default=default, ensure_ascii=False))