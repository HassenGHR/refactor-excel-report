#!/usr/bin/env python3
"""
tp195_extract.py — extract a TP-195 (AIN T'SILA, ENTP rig, AT-NN wells)
Daily Workover Report into the standard dict shape.

Source layout
-------------
~122 rows × 24 cols, English/French.  Same SONATRACH PRODUCTION DIVISION
template family as TP-182 but with cells split into separate label/value
pairs rather than combined "LABEL: value" strings.

Header (rows 5-7): label in col B/D/G/I/K/M/O, value in the next col.
Mud panel (rows 11-29 cols J-O): left side J=label, L=value; right side
M=label, N/O=value.
Mud chemicals (rows 47-54): left B/E/F/G/H, right I/L/M/O/P.
Survey data (rows 65-69): MD/INC/AZI/TVD/N/E/DLS/VS columns.
Personnel (rows 72-75): left B/D/G, right J/L/N.
Operations (rows 78-91): FROM/TO/HRS/DESC/BILL/COMPANY headers at row 77.
Text sections (rows 103-113): label col B, value col D.

Distinguishing markers from other SONATRACH-family templates:
    - "OFFICE REP" label (TP-182 has "SUPERINTANDANT" instead)
    - "REPRISE DATE" label
    - "NEXT BOP TEST" (no LAST BOP)
"""
from __future__ import annotations
import re
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _clean(v) -> str:
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _float(v, default=0.0) -> float:
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    s = _clean(v).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "None"): return default
    s = re.sub(r"[a-zA-Zé°%/³]+$", "", s).strip()
    try: return float(s)
    except ValueError: return default


def _int(v, default=0) -> int:
    return int(_float(v, float(default)))


def _date_parse(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date_type): return v
    s = _clean(v)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d,%m,%y", "%d/%m/%y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None


def _time_parse(v):
    if v is None: return None
    if isinstance(v, time): return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        # 24h durations are represented as 1-day 0:00:00 (=86400 sec)
        if total == 86400:
            return time(0, 0)
        return time((total // 3600) % 24, (total % 3600) // 60)
    s = _clean(v)
    if not s or s in ("-", "None", "/"): return None
    if s in ("24:00", "24:00:00"): return time(0, 0)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try: return datetime.strptime(s, fmt).time()
        except ValueError: continue
    return None


def _duration_hours(v) -> float:
    """Decode an HRS cell — could be number, timedelta, time, or '1 day 0:00:00' string."""
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, timedelta):
        return v.total_seconds() / 3600.0
    if isinstance(v, time):
        return v.hour + v.minute / 60.0
    if isinstance(v, datetime):
        # Excel epoch zero (1900-01-01 00:00) means 24-hour duration
        if v.year == 1900 and v.month == 1 and v.day == 1:
            return v.hour + v.minute / 60.0
        return 0.0
    s = _clean(v)
    m = re.match(r"^(\d+)\s+day", s)
    if m:
        # "1 day, 0:00:00" → 24h
        days = int(m.group(1))
        rest = re.sub(r"^\d+\s+day,?\s*", "", s)
        t = _time_parse(rest)
        return days * 24 + (t.hour + t.minute/60.0 if t else 0)
    t = _time_parse(v)
    if t: return t.hour + t.minute / 60.0
    return _float(v)


def _build_merged_lookup(ws):
    lookup = {}
    for mr in ws.merged_cells.ranges:
        av = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = av
    return lookup


def _cell(ws, r, c, lookup):
    v = ws.cell(r, c).value
    return v if v is not None else lookup.get((r, c))


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------
def parse_tp195(source: Union[Path, str, BytesIO]) -> dict:
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    ws = wb.active
    L = _build_merged_lookup(ws)

    # =====================================================================
    # HEADER (rows 5-7) — label/value pairs in adjacent cells
    # =====================================================================
    header = {}

    # Row 5
    header["well_name"]      = _clean(_cell(ws, 5, 3,  L))     # C5
    header["date"]           = _date_parse(_cell(ws, 5, 5,  L))  # E5
    header["well_md"]        = _float(_cell(ws, 5, 8,  L))     # H5  Total MD
    header["tvd"]            = _float(_cell(ws, 5, 10, L))     # J5
    header["formation_top"]  = _clean(_cell(ws, 5, 12, L))     # L5
    header["reprise_date"]   = _date_parse(_cell(ws, 5, 14, L))  # N5
    header["day_number"]     = _int(_cell(ws, 5, 16, L))       # P5

    # Row 6  — SUPERVISOR / RIG / FIELD / OFFICE REP / ACC FREE / DAILY NR
    # Per the user's convention: SUPERVISOR + OFFICE REP are the two
    # 12-hour-shift supervisors.  Slot them into supervisor + superintendent.
    header["supervisor"]          = _clean(_cell(ws, 6, 3,  L))   # C6  SUPERVISOR
    header["rig_name"]            = _clean(_cell(ws, 6, 6,  L))   # F6
    header["field_name"]          = _clean(_cell(ws, 6, 9,  L))   # I6
    header["superintendent"]      = _clean(_cell(ws, 6, 11, L))   # K6  OFFICE REP -> second supervisor
    header["accident_free_days"]  = _int(_cell(ws, 6, 14, L))     # N6
    header["daily_npt"]           = _duration_hours(_cell(ws, 6, 16, L))  # P6

    # Row 7
    header["last_formation_test"] = _clean(_cell(ws, 7, 3,  L))   # C7
    # KOP / LAST CSG SHOE: only depth, no size info in this template.
    csg_depth = _cell(ws, 7, 7,  L)                                # G7
    if csg_depth is not None:
        try:
            d = int(float(str(csg_depth)))
            header["top_shoe"] = f"@ {d}m"                          # frontend "Last CSG SHOE"
        except (ValueError, TypeError):
            header["top_shoe"] = f"@ {_clean(csg_depth)}"
    header["last_tol"]            = _clean(_cell(ws, 7, 9,  L))    # I7  Last TOL
    # NEXT BOP TEST (not LAST) — store separately, don't put in bop_test
    next_bop = _date_parse(_cell(ws, 7, 11, L))                    # K7
    if next_bop:
        header["next_bop_test"] = next_bop
    header["last_safety_meeting"] = _date_parse(_cell(ws, 7, 14, L))  # N7
    header["cum_npt"]             = _duration_hours(_cell(ws, 7, 16, L))  # P7

    # Workover reason — row 8 (or 9), label in B, value in C
    wr = _cell(ws, 8, 3, L) or _cell(ws, 9, 3, L)
    if wr:
        header["well_objective"] = _clean(wr)

    # =====================================================================
    # BIT DATA (rows 11-27 cols B label, E value) — extract key fields
    # =====================================================================
    bit_data = {}
    bit_labels = [
        (13, "bit_size"),
        (14, "bit_manufacture"),
        (15, "bit_type"),
        (16, "bit_iadc"),
        (17, "bit_serial"),
        (18, "bit_jets"),
        (19, "bit_tfa"),
        (20, "bit_depth_in"),
    ]
    for row, key in bit_labels:
        v = _cell(ws, row, 5, L)        # col E
        if v is not None:
            bit_data[key] = _clean(v)
    if bit_data:
        header["bit_data"] = " | ".join(f"{k}={v}" for k, v in bit_data.items())

    # =====================================================================
    # LAST BHA (rows 13-29 cols F desc / H OD / I length)
    # =====================================================================
    bha_parts = []
    for row in range(13, 30):
        name = _clean(_cell(ws, row, 6, L) or "")     # F
        if not name or name.upper() == "BHA TOTAL":
            continue
        od     = _clean(_cell(ws, row, 8, L) or "")   # H
        length = _cell(ws, row, 9, L)                 # I
        if length is not None:
            try: length = f"{float(length):.2f}m"
            except: length = _clean(length)
        parts = [name]
        if od: parts.append(od)
        if length: parts.append(str(length))
        bha_parts.append(" / ".join(parts))
    if bha_parts:
        header["bha_details"] = " | ".join(bha_parts)

    # =====================================================================
    # MUD CHECKS (rows 11-29; J=label L=value left | M=label N/O=value right)
    # =====================================================================
    mud_checks = {}
    mud_type_raw = _cell(ws, 11, 13, L) or _cell(ws, 11, 14, L)  # M11 / N11
    if mud_type_raw:
        s = _clean(mud_type_raw)
        if s.upper() not in ("MUD", "TYPE", "TYPE1/TYPE2"):
            mud_checks["mud_type"] = s

    left_mud = [
        (13, "density"),
        (14, "depth"),
        (15, "fl_tmp"),
        (16, "fun_vis"),
        (17, "pv"),
        (18, "yp"),
        (19, "gel10sec"),
        (20, "gel10m"),
        # row 21 = HPHT FILTRAT
        (21, "hpht_fl"),
        (22, "lgs"),
        # (23, "lgs_caco3"),
        (24, "hgs"),
        # (25, "asgs"), (26, "asgs"),
        (27, "solid"),    # CORRECTED SOLID
    ]
    for row, key in left_mud:
        v = _cell(ws, row, 12, L)        # col L
        if v is None: continue
        try: mud_checks[key] = _float(v)
        except: pass

    right_mud = [
        # rows 13-26 right side mostly empty in this report, but capture anyway
        (13, "pom"),
        (14, "solids"),
        (15, "oil"),
        (16, "h2o"),
        # row 17 OIL/WATER (ratio text), 18-26 various
        (21, "es"),
        # (22, "brine"), (23, "caco3"),
        # row 27 = LSYS — value at col O
    ]
    for row, key in right_mud:
        v = _cell(ws, row, 14, L) or _cell(ws, row, 15, L)
        if v is None: continue
        try: mud_checks[key] = _float(v)
        except: pass

    # LSYS at row 27 col O
    lsys = _cell(ws, 27, 15, L)
    if lsys is not None:
        try: mud_checks["lsys"] = _float(lsys)
        except: pass

    # =====================================================================
    # MUD VOLUMES
    # =====================================================================
    mud_volume = {}
    # Well Vol at r28 col O
    well_v = _cell(ws, 28, 15, L)
    if well_v is not None: mud_volume["string_volume"] = _float(well_v)
    # OBM Surface Vol at r28 col L (left side)
    obm_surface = _cell(ws, 28, 12, L)
    if obm_surface is not None: mud_volume["pits_volume"] = _float(obm_surface)
    # Losses at r29 col L
    losses = _cell(ws, 29, 12, L)
    if losses is not None:
        try:
            mud_volume["surface_loss"] = _float(losses)
        except:
            pass
    # Gas-oil Vol at r29 col O
    gas_oil = _cell(ws, 29, 15, L)
    if gas_oil is not None:
        try:
            mud_volume["gas_oil_volume"] = _float(gas_oil)
        except:
            pass
    if mud_volume:
        mud_volume["total_volume"] = sum(
            v for v in (mud_volume.get("string_volume"), mud_volume.get("pits_volume"))
            if v is not None
        )

    # =====================================================================
    # MUD CHEMICAL USAGE  (rows 47-54)
    # Left block:  col B item, col E units, col F rec'd, col G used, col H end
    # Right block: col I item, col L units, col M rec'd, col O used, col P end
    # =====================================================================
    chemicals = []
    for row in range(47, 55):
        # LEFT
        item_l = _clean(_cell(ws, row, 2, L) or "")
        if item_l:
            chemicals.append({
                "item": item_l,
                "units": _clean(_cell(ws, row, 5, L) or ""),
                "received": _clean(_cell(ws, row, 6, L) if _cell(ws, row, 6, L) is not None else ""),
                "used":     _clean(_cell(ws, row, 7, L) if _cell(ws, row, 7, L) is not None else ""),
                "on_loc":   _clean(_cell(ws, row, 8, L) if _cell(ws, row, 8, L) is not None else "0"),
            })
        # RIGHT
        item_r = _clean(_cell(ws, row, 9, L) or "")
        if item_r:
            chemicals.append({
                "item": item_r,
                "units": _clean(_cell(ws, row, 12, L) or ""),
                "received": _clean(_cell(ws, row, 13, L) if _cell(ws, row, 13, L) is not None else ""),
                "used":     _clean(_cell(ws, row, 15, L) if _cell(ws, row, 15, L) is not None else ""),
                "on_loc":   _clean(_cell(ws, row, 16, L) if _cell(ws, row, 16, L) is not None else "0"),
            })

    # =====================================================================
    # SURVEY DATA (rows 68-69)
    # Header r65: B=COMMENTS C=MD D=INCL F=AZI H=TVD J=NORTH L=EAST N=DLS P=VS
    # =====================================================================
    survey_data = []
    for row in (68, 69):
        comment = _clean(_cell(ws, row, 2, L) or "")
        md = _cell(ws, row, 3, L)
        if not comment or md is None:
            continue
        survey_data.append({
            "md":  _float(md),
            "inc": _float(_cell(ws, row, 4, L)),
            "azi": _float(_cell(ws, row, 6, L)),
            "tvd": _float(_cell(ws, row, 8, L)),
            "n_s": _float(_cell(ws, row, 10, L)),
            "e_w": _float(_cell(ws, row, 12, L)),
            "dls": _float(_cell(ws, row, 14, L)),
            "vs":  _float(_cell(ws, row, 16, L)),
        })

    # =====================================================================
    # PERSONNEL (rows 72-75)
    # Left  : col B company, col D number, col E hours, col G names
    # Right : col J company, col L number, col M hours, col N names
    # =====================================================================
    personnel = []
    for row in range(72, 76):
        # LEFT
        company_l = _clean(_cell(ws, row, 2, L) or "")
        if company_l:
            count_l = _cell(ws, row, 4, L)
            names_l = _clean(_cell(ws, row, 7, L) or "")
            if "WATER TRUCK" in company_l.upper() and count_l is not None:
                header["water_truck"] = _int(count_l)
            personnel.append({
                "company": company_l,
                "number":  _int(count_l, 0),
                "hours":   _clean(_cell(ws, row, 5, L) or ""),
                "names":   names_l,
            })
        # RIGHT
        company_r = _clean(_cell(ws, row, 10, L) or "")
        if company_r:
            count_r = _cell(ws, row, 12, L)
            names_r = _clean(_cell(ws, row, 14, L) or "")
            personnel.append({
                "company": company_r,
                "number":  _int(count_r, 0),
                "hours":   _clean(_cell(ws, row, 13, L) or ""),
                "names":   names_r,
            })

    # =====================================================================
    # OPERATIONS  (rows 78+)
    # Header r77: B=FROM C=TO D=HRS E=DESCRIPTION O=BILL P=COMPANY
    # =====================================================================
    activities = []
    for row in range(78, 100):
        start = _cell(ws, row, 2, L)
        end   = _cell(ws, row, 3, L)
        hrs   = _cell(ws, row, 4, L)
        desc  = _clean(_cell(ws, row, 5, L) or "")
        bill_raw = _clean(_cell(ws, row, 15, L) or "")
        comp  = _clean(_cell(ws, row, 16, L) or "")

        if not bill_raw and not desc:
            continue
        # Real ops have a BILL code
        if not bill_raw:
            continue
        # Skip summary lines
        if any(t in desc.upper() for t in ("T1=", "T2=", "T3=")):
            continue

        # Normalize bill code: this template stores values like "1,05xT1"
        # (meaning "1.05 multiplier × T1 hourly rate").  The insert function
        # looks for the strict pattern ^T(\d+)$ to set tValue, so we strip
        # any leading multiplier and keep only the T-letter.  The
        # multiplier itself isn't used by the cost calculator — the rig's
        # ValueRig/ValueRigV2 tariff already encodes the applicable rate.
        m = re.search(r"T\s*(\d+)", bill_raw, re.IGNORECASE)
        bill = f"T{m.group(1)}" if m else bill_raw

        start_t = _time_parse(start)
        end_t   = _time_parse(end)
        hours = _duration_hours(hrs)
        if hours == 0.0 and start_t and end_t:
            sm = start_t.hour * 60 + start_t.minute
            em = end_t.hour * 60 + end_t.minute
            if em == sm: hours = 24.0
            elif em > sm: hours = (em - sm) / 60.0
            else: hours = (em + 1440 - sm) / 60.0

        activities.append({
            "start_time": start_t,
            "end_time":   end_t,
            "hours":      hours,
            "phase_name": "",
            "code": "", "sub": "",
            "description": desc,
            "start_md": 0, "end_md": 0,
            "npt": 0, "npt_detail": "",
            "npt_company": "", "op_company": comp,
            "bill": bill,
        })

    # =====================================================================
    # TEXT SECTIONS  (rows 103-113, col B label, col D value)
    # =====================================================================
    text_sections = {}
    for row in range(103, 115):
        v = _cell(ws, row, 2, L)
        if v is None: continue
        text = _clean(v)
        if not text: continue
        upper = text.upper()
        val = _cell(ws, row, 4, L)
        val_clean = _clean(val) if val else ""
        if not val_clean or val_clean.upper() == upper:
            continue
        if upper.startswith("ACTUEL OPERATIONS") or "ACTUEL OPERATIONS" in upper[:30]:
            text_sections["current_operation"] = val_clean
            text_sections["day_summary"]       = val_clean
        elif upper.startswith("PLAN OPERATION"):
            text_sections["plan_operations"] = val_clean
        elif "REQUIREMENT" in upper or "RENTAL EQUIPMENT" in upper:
            text_sections["requirements"] = val_clean
        elif upper.startswith("REMARK"):
            # Skip "/" placeholder remarks
            if val_clean != "/":
                text_sections["remarks"] = val_clean

    # =====================================================================
    # SAFETY
    # =====================================================================
    safety = {}
    if header.get("accident_free_days") is not None:
        safety["accident_free_days"] = header["accident_free_days"]
    if header.get("water_truck") is not None:
        safety["water_truck"] = header["water_truck"]

    wb.close()

    return {
        "header": header,
        "activities": activities,
        "text_sections": text_sections,
        "mud_checks": mud_checks,
        "mud_volume": mud_volume,
        "mud_chemical_usage": chemicals,
        "personnel_data": personnel,
        "pumps": [],
        "well_location": {},
        "survey_data": survey_data,
        "safety": safety,
        "tarif_totals": {},
    }


# Drop-in compat
parse_daily_excel_report = parse_tp195


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Usage: tp195_extract.py SOURCE.xlsx")
    data = parse_tp195(Path(sys.argv[1]))

    def default(o):
        if isinstance(o, (date_type, datetime)): return o.isoformat()
        if isinstance(o, time): return o.strftime("%H:%M:%S")
        if isinstance(o, timedelta): return o.total_seconds()
        return str(o)
    print(json.dumps(data, indent=2, default=default, ensure_ascii=False))