#!/usr/bin/env python3
"""
parse_source.py — auto-detect rig-report format and dispatch to the
appropriate extractor.

Accepts both Excel sources (.xlsx — most rigs) and PDF sources (some rigs
deliver their daily reports as PDF instead of Excel). The file type is
sniffed from the first few bytes (PDF starts with %PDF-, xlsx with PK\\x03)
so callers don't need to declare the format up front.

Usage (programmatic):
    from parse_source import parse_source
    data = parse_source(Path("report.xlsx"))   # Excel
    data = parse_source(Path("report.pdf"))    # PDF
    data = parse_source(BytesIO(file_bytes))   # any in-memory source

Usage (CLI):
    python parse_source.py report.xlsx           # prints detected format
    python parse_source.py report.pdf --json     # prints the extracted dict
"""
from __future__ import annotations
import re
import sys
from pathlib import Path
from typing import Union
from io import BytesIO

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# File-type sniffing — by magic bytes, not by extension
# ---------------------------------------------------------------------------
def _read_head(source) -> bytes:
    """Read the first 8 bytes of a source without consuming it."""
    if isinstance(source, (str, Path)):
        with open(source, "rb") as f:
            return f.read(8)
    # BytesIO or similar: remember position, peek, restore
    pos = source.tell()
    head = source.read(8)
    source.seek(pos)
    return head


def _sniff_kind(source) -> str:
    """Return 'pdf', 'xlsx', or 'unknown' based on the file's magic bytes."""
    head = _read_head(source)
    if head.startswith(b"%PDF-"):
        return "pdf"
    # xlsx is a zip archive — starts with PK\x03\x04
    if head.startswith(b"PK\x03\x04"):
        return "xlsx"
    return "unknown"


# ---------------------------------------------------------------------------
# Excel format detection
# ---------------------------------------------------------------------------
def _detect_format_xlsx(source) -> str:
    """Peek at an Excel source and return the rig template key."""
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True, read_only=True)
    else:
        wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb.active

    # Scan the first 12 rows × 28 cols for marker strings
    markers = []
    for row in ws.iter_rows(min_row=1, max_row=12, max_col=28, values_only=True):
        for v in row:
            if v is not None:
                markers.append(str(v).upper())
    blob = " || ".join(markers)
    wb.close()

    # TP-195 — SONATRACH AIN T'SILA format.  Same template family as TP-182
    # (English, "DAILY DRILLING REPORT" title) but uses split label/value
    # cells, "OFFICE REP" instead of "SUPERINTANDANT", and only "NEXT BOP
    # TEST" (no LAST BOP).  Must be checked BEFORE TP-182 since both share
    # the SONATRACH PRODUCTION DIVISION title.
    if "OFFICE REP" in blob:
        return "tp195"

    # TP-182 — SONATRACH PRODUCTION DIVISION Daily Drilling Report format
    # (English, "SUPERINTANDANT" misspelling, has "WORKOVER REASON")
    if "SUPERINTANDANT" in blob:        # specific to TP-182 template
        return "tp182"
    if "SONATRACH PRODUCTION DIVISION" in blob and "DAILY DRILLING REPORT" in blob:
        return "tp182"

    # ENAFOR ENF#04 — French workover format (Haoud Berkaoui, DDNH wells).
    # Title "RAPPORT JOURNALIER DE WORKOVER" (note "DE" — distinguishes from
    # TP-179's "DU WORK OVER").  Must be checked BEFORE TP-179 since both
    # share "RAPPORT JOURNALIER" + "WORK".
    if "HAOUD BERKAOUI" in blob or "RAPPORT JOURNALIER DE WORKOVER" in blob:
        return "enf04"

    # GW-series rigs (GWDC operator, RBL wells, REB field) — French workover
    # format but with distinct AVANCEMENT/OUTILS/USURE/PARAMETRES layout.
    # Must be checked BEFORE the generic TP-179 catch-all below since GW29
    # also has "RAPPORT JOURNALIER" + "WORK".
    if "RAPPORT JOURNALIER" in blob and ("AVANCEMENT" in blob
                                         or "PARAMETRES" in blob
                                         or re.search(r"\bGW\s?\d{2}\b", blob)):
        return "gw29"

    # TP-173 — same template family as TP-179 (RAPPORT JOURNALIER DU WORK
    # OVER, ADRAR region) but with different cell positions: "APPAREIL:"
    # instead of "RIG :", "DERNIER TUBAGE:" combined string for casing.
    # Must be checked BEFORE TP-179 since both share the title.
    if "TP-173" in blob or "DERNIER TUBAGE" in blob:
        return "tp173"

    # TP-179 — French workover format (ENTP rigs)
    if "RAPPORT JOURNALIER" in blob and "WORK" in blob:
        return "tp179"

    # ENAFOR DDR (ENF#NN rigs) — distinguishing markers
    if "DAILY DRILLING REPORT" in blob and ("ENF#" in blob or "ENF #" in blob):
        return "enf"

    # Generic drilling — try enf as a fallback
    if "DAILY DRILLING REPORT" in blob:
        return "enf"

    return "unknown"


# ---------------------------------------------------------------------------
# PDF format detection
# ---------------------------------------------------------------------------
def _detect_format_pdf(source) -> str:
    """Peek at a PDF source and return the rig template key.

    Reads the first 2 pages and scans for distinguishing marker phrases.
    pdfplumber is imported lazily so the module still loads on systems
    without it when only Excel sources are used.
    """
    import pdfplumber

    if isinstance(source, BytesIO):
        # pdfplumber consumes BytesIO; rewind it for the caller after detection
        pos = source.tell()
        try:
            with pdfplumber.open(source) as pdf:
                pages_to_scan = pdf.pages[:2]
                blob = " || ".join(
                    (p.extract_text() or "").upper() for p in pages_to_scan
                )
        finally:
            source.seek(pos)
    else:
        with pdfplumber.open(source) as pdf:
            pages_to_scan = pdf.pages[:2]
            blob = " || ".join(
                (p.extract_text() or "").upper() for p in pages_to_scan
            )

    # ENF#34 — Gassi-Touil workover PDF (this is the first PDF format we
    # support).  Distinguishing markers: "GASSI" + "RAPPORT JOURNALIER DE
    # WORK OVER" (note the space in "WORK OVER" — TP-179 uses "WORKOVER").
    if "GASSI" in blob and "RAPPORT JOURNALIER" in blob and "WORK" in blob:
        return "enf34_pdf"
    if "DIRECTION RÉGIONALE GASSI" in blob or "GASSI-TOUIL" in blob:
        return "enf34_pdf"

    return "unknown"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def _detect_format(source) -> str:
    """Single entry point for format detection.  Returns a rig key
    like 'enf', 'tp182', 'enf34_pdf', etc.  Returns 'unknown' if the file
    type or layout couldn't be identified."""
    kind = _sniff_kind(source)
    if kind == "pdf":
        return _detect_format_pdf(source)
    if kind == "xlsx":
        return _detect_format_xlsx(source)
    return "unknown"


def parse_source(source: Union[Path, str, BytesIO]) -> dict:
    """Detect the source format and call the right extractor.

    Accepts both Excel (.xlsx) and PDF inputs.  Returns the standard
    dict shape produced by all extractors (header / activities /
    text_sections / mud_checks / mud_volume / mud_chemical_usage /
    personnel_data / pumps / well_location / survey_data / safety /
    tarif_totals) plus a "_meta" key with the source format and kind.
    """
    fmt = _detect_format(source)

    # Excel-backed extractors
    if fmt == "enf":
        from extractors.enf17_extract import parse_enf17 as parse_ddr
        data = parse_ddr(source)
    elif fmt == "tp179":
        from extractors.tp179_extract import parse_tp179
        data = parse_tp179(source)
    elif fmt == "tp173":
        from extractors.tp173_extract import parse_tp173
        data = parse_tp173(source)
    elif fmt == "tp182":
        from extractors.tp182_extract import parse_tp182
        data = parse_tp182(source)
    elif fmt == "tp195":
        from extractors.tp195_extract import parse_tp195
        data = parse_tp195(source)
    elif fmt == "gw29":
        from extractors.gw29_extract import parse_gw29
        data = parse_gw29(source)
    elif fmt == "enf04":
        from extractors.enf04_extract import parse_enf04
        data = parse_enf04(source)

    # PDF-backed extractors
    elif fmt == "enf34_pdf":
        from extractors.enf34_pdf_extract import parse_enf34_pdf
        data = parse_enf34_pdf(source)

    else:
        raise ValueError(
            f"Unrecognised report format. Markers in the file did not match "
            f"any known rig layout. Add a new extractor module and register "
            f"it in parse_source._detect_format_xlsx() or "
            f"_detect_format_pdf()."
        )

    data.setdefault("_meta", {})["source_format"] = fmt
    data["_meta"]["source_kind"] = _sniff_kind(source)
    return data


def main(argv=None) -> int:
    import argparse, json
    from datetime import date as date_type, datetime, time, timedelta

    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("source", type=Path)
    p.add_argument("--json", action="store_true", help="dump the extracted dict")
    args = p.parse_args(argv)

    if not args.source.exists():
        sys.exit(f"ERROR: source not found: {args.source}")

    fmt = _detect_format(args.source)
    print(f"Detected format: {fmt}", file=sys.stderr)

    if args.json:
        data = parse_source(args.source)
        def default(o):
            if isinstance(o, (date_type, datetime)): return o.isoformat()
            if isinstance(o, time): return o.strftime("%H:%M:%S")
            if isinstance(o, timedelta): return o.total_seconds()
            return str(o)
        print(json.dumps(data, indent=2, default=default))
    return 0


if __name__ == "__main__":
    sys.exit(main())