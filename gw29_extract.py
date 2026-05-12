#!/usr/bin/env python3
"""
gw29_extract.py — extract a GW29 (GWDC operator, RBL wells, REB field)
Daily Workover Report into the standard dict shape.

Source layout
-------------
~57 rows × 30 cols, French. Title at H1: "Rapport journalier de Work - Over".
Distinguishing markers from other "RAPPORT JOURNALIER" formats:
    - AVANCEMENT / OUTILS / USURE / PARAMETRES section headers at row 3
    - GW-series rig name in I2
    - GWDC contractor in the personnel column
Header at rows 1-2 (cell-pair "Label + value" style).
Mud panel interleaved at cols W-Z (W=label X=value | Y=label Z=value).
Operations table at rows 17-23 (A/B/C/M/N).
After-midnight at r34-r35 (C=text, A=start, B=end).
Personnel at rows 33-39 (W=role, Z=count) — role counts only, no individual names.
"""
from __future__ import annotations
import re
from datetime import datetime, time, date as date_type, timedelta
from io import BytesIO
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers (same shape as the other extractors)
# ---------------------------------------------------------------------------
def _clean(v) -> str:
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _float(v, default=0.0) -> float:
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    s = _clean(v).replace(",", ".").replace(" ", "")
    if s in ("", "-", "/", "None"): return default
    s = re.sub(r"[a-zA-Zé°%/³]+$", "", s).strip()
    try: return float(s)
    except ValueError: return default


def _int(v, default=0) -> int:
    return int(_float(v, float(default)))


def _date_parse(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date_type): return v
    s = _clean(v)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d,%m,%y", "%d/%m/%y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None


def _time_parse(v):
    if v is None: return None
    if isinstance(v, time): return v
    if isinstance(v, datetime):
        # Excel epoch zero (1900-01-01) means midnight / end-of-day
        return v.time()
    if isinstance(v, timedelta):
        total = int(v.total_seconds())
        return time((total // 3600) % 24, (total % 3600) // 60)
    s = _clean(v)
    if not s or s in ("-", "None"): return None
    if s in ("24:00", "24:00:00"): return time(0, 0)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try: return datetime.strptime(s, fmt).time()
        except ValueError: continue
    return None


def _build_merged_lookup(ws):
    lookup = {}
    for mr in ws.merged_cells.ranges:
        av = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = av
    return lookup


def _cell(ws, r, c, lookup):
    v = ws.cell(r, c).value
    return v if v is not None else lookup.get((r, c))


def _format_csg(label: str, depth_val) -> str:
    """Combine 'Last Csg 7"' + 3760 -> '7" @ 3760m'."""
    label_clean = _clean(label)
    # Strip "Last Csg" / "Top Liner" prefix from the size description
    size = re.sub(r"^\s*(?:Last\s+Csg|Top\s+Liner|Last\s+Lnr)\s*",
                  "", label_clean, flags=re.IGNORECASE).strip()
    if depth_val is None:
        return size
    try:
        depth = int(float(str(depth_val)))
        return f"{size} @ {depth}m"
    except (ValueError, TypeError):
        return f"{size} @ {_clean(depth_val)}"


# ---------------------------------------------------------------------------
# Main extractor
# ---------------------------------------------------------------------------
def parse_gw29(source: Union[Path, str, BytesIO]) -> dict:
    if isinstance(source, (str, Path)):
        wb = load_workbook(source, data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    ws = wb.active
    L = _build_merged_lookup(ws)

    # =====================================================================
    # HEADER (rows 1-2)
    # =====================================================================
    header = {}
    header["well_name"]    = _clean(_cell(ws, 2, 2,  L))     # B2 "RBL-03"
    header["field_name"]   = _clean(_cell(ws, 2, 5,  L))     # E2 "REB"
    header["rig_name"]     = _clean(_cell(ws, 2, 9,  L))     # I2 "GW29"
    header["date"]         = _date_parse(_cell(ws, 1, 21, L))  # U1
    header["day_number"]   = _int(_cell(ws, 1, 26, L))       # Z1

    # Total depth (TD) — label at S2, value at U2
    header["well_md"]      = _float(_cell(ws, 2, 21, L))     # U2

    # Casing (composite label/value): K2 "Last Csg 7\"" + L2 3760
    csg_label = _cell(ws, 2, 11, L)                          # K2
    csg_depth = _cell(ws, 2, 12, L)                          # L2
    if csg_label or csg_depth:
        formatted = _format_csg(str(csg_label or ""), csg_depth)
        # Per the user's mapping convention:
        #   "Last Csg ..."  in the source     -> header.top_shoe -> DB.lastCSNlnrSHOE -> frontend "Last CSG SHOE"
        # So we put the casing shoe info there (NOT in last_csg_shoe).
        header["top_shoe"] = formatted

    # Top Liner: N2 "Top Liner 4\"1/2" + P2 3658
    liner_label = _cell(ws, 2, 14, L)                        # N2
    liner_depth = _cell(ws, 2, 16, L)                        # P2
    if liner_label or liner_depth:
        # Top Liner in the source -> header.last_csg_top -> DB.lastCSNlnrTOP
        # -> frontend "Last LNR TOP"
        size = re.sub(r"^\s*Top\s+Liner\s*", "",
                      _clean(liner_label or ""), flags=re.IGNORECASE).strip()
        if liner_depth is not None:
            try:
                d = int(float(str(liner_depth)))
                header["last_csg_top"] = f"{size} @ {d}m" if size else f"@ {d}m"
            except (ValueError, TypeError):
                header["last_csg_top"] = f"{size} @ {_clean(liner_depth)}".strip()
        else:
            header["last_csg_top"] = size

    # BOP test — earliest date in O36-O39 (all components tested same day usually)
    for row in range(36, 40):
        d = _date_parse(_cell(ws, row, 15, L))
        if d:
            header["bop_test"] = d
            break

    # =====================================================================
    # MUD CHECKS  (rows 6-16, cols W-Z — interleaved label/value pairs)
    # =====================================================================
    mud_checks = {}
    mud_type = _cell(ws, 3, 25, L) or _cell(ws, 3, 23, L)    # Y3 "OBM" (or W3 type label)
    if mud_type:
        s = _clean(mud_type)
        if s.upper() not in ("MUD", "TYPE", "FAB BOUE"):
            mud_checks["mud_type"] = s

    # Pair list: (row, our_dict_key, side)
    # side 'L' = label at col W (23), value at col X (24)
    # side 'R' = label at col Y (25), value at col Z (26)
    mud_panel = [
        (10, "density",  "L"),    # Densité           1.38
        (10, "pv",       "R"),    # Plast Vis         30
        (11, "fun_vis",  "L"),    # V. Marsh          90
        (11, "yp",       "R"),    # Yield-p           15
        (12, "apl_fl",   "L"),    # Filtrat           5
        # (12, "lsys",    "R"),    # LSYS (not in our schema)
        (13, "es",       "L"),    # E stability       1200
        (13, "gel10sec", "R"),    # Gel 0/10 split below
        (14, "solid",    "L"),    # Solide (%)        16
        (14, "ph",       "R"),    # PH
        # row 15 W/X "Oil/w/ratio" — text, handled separately
    ]
    for row, key, side in mud_panel:
        col = 24 if side == "L" else 26                       # X or Z
        v = _cell(ws, row, col, L)
        if v is None: continue
        if key == "gel10sec" and isinstance(v, str) and "/" in v:
            # "8/12" -> gel10sec=8, gel10m=12
            try:
                a, b = v.split("/", 1)
                mud_checks["gel10sec"] = _float(a)
                mud_checks["gel10m"]   = _float(b)
            except ValueError:
                pass
        else:
            try:
                mud_checks[key] = _float(v)
            except Exception:
                pass

    # Oil/water ratio is a text "90/10"
    owr = _cell(ws, 15, 24, L)
    if owr:
        s = _clean(owr)
        if s:
            mud_checks["oil_water_ratio"] = s

    # =====================================================================
    # MUD VOLUMES (rows 6-9 col Z, label in col W)
    # =====================================================================
    mud_volume = {}
    string_v = _cell(ws, 8, 26, L)      # Z8 = Volume puits
    surface_v = _cell(ws, 9, 26, L)     # Z9 = Totale surface
    surface_loss = _cell(ws, 6, 26, L)  # Z6 = Perte de surface
    if string_v is not None:  mud_volume["string_volume"] = _float(string_v)
    if surface_v is not None: mud_volume["pits_volume"]  = _float(surface_v)
    if surface_loss is not None:
        mud_volume["surface_loss"] = _float(surface_loss)
    if mud_volume:
        mud_volume["total_volume"] = sum(
            v for v in (mud_volume.get("string_volume"),
                        mud_volume.get("pits_volume"))
            if v is not None
        )

    # =====================================================================
    # MUD CHEMICAL USAGE (rows 19-23 col W = item, cols Y/Z = used/stock)
    # =====================================================================
    chemicals = []
    for row in range(19, 25):
        item = _clean(_cell(ws, row, 23, L) or "")        # col W
        if not item or "PRODUITS" in item.upper():
            continue
        used  = _cell(ws, row, 25, L)                      # Y
        stock = _cell(ws, row, 26, L)                      # Z
        chemicals.append({
            "item": item,
            "units": "",
            "received": "",
            "used":   _clean(used  if used  is not None else ""),
            "on_loc": _clean(stock if stock is not None else ""),
        })

    # =====================================================================
    # PERSONNEL (rows 33-39 cols W=role, Z=count) — role counts only,
    # no individual names in this format.  The "Maitre d'œuvres SH-DP" row
    # would be the supervisor's role line but the actual name isn't on the
    # sheet, so we leave header.supervisor / header.superintendent unset.
    # =====================================================================
    personnel = []
    for row in range(33, 40):
        role  = _clean(_cell(ws, row, 23, L) or "")        # W
        count = _cell(ws, row, 26, L)                      # Z
        if not role or "PERSONNEL" in role.upper():
            continue
        if count is None and not role:
            continue
        personnel.append({
            "company": role,
            "number":  _int(count, 0),
            "hours":   "",
            "names":   "",
        })

    # =====================================================================
    # OPERATIONS (rows 17-30)
    # Header at row 17: A=DE, B=A, C=Opérations, M=Code, N=Heures
    # =====================================================================
    activities = []
    for row in range(18, 31):
        start = _cell(ws, row, 1, L)            # A
        end   = _cell(ws, row, 2, L)            # B
        desc  = _clean(_cell(ws, row, 3, L) or "")
        bill  = _clean(_cell(ws, row, 13, L) or "")
        hrs   = _cell(ws, row, 14, L)

        if not bill and not desc and start is None:
            continue
        if not bill:
            continue
        # Skip total/summary lines
        if "TOTAL" in (bill.upper() + desc.upper()):
            continue

        start_t = _time_parse(start)
        end_t   = _time_parse(end)
        hours = 0.0
        if isinstance(hrs, (int, float)):
            hours = float(hrs)
        elif start_t and end_t:
            sm = start_t.hour * 60 + start_t.minute
            em = end_t.hour * 60 + end_t.minute
            if em == sm: hours = 24.0
            elif em > sm: hours = (em - sm) / 60.0
            else: hours = (em + 1440 - sm) / 60.0

        activities.append({
            "start_time": start_t,
            "end_time":   end_t,
            "hours":      hours,
            "phase_name": "",
            "code": "", "sub": "",
            "description": desc,
            "start_md": 0, "end_md": 0,
            "npt": 0, "npt_detail": "",
            "npt_company": "", "op_company": "",
            "bill": bill,
        })

    # =====================================================================
    # TEXT SECTIONS
    #   after_midnight   : r34 C "Après minuit:" label, r35 C value
    #   current_operation: r40 A label "Situation au rapport @ 06h00",
    #                      r40 D value
    #   plan_operations  : r41 A label "Programme prévu", value if any
    # =====================================================================
    text_sections = {}

    # After midnight
    am_label = _cell(ws, 34, 3, L) or ""
    if "APRÈS MINUIT" in str(am_label).upper() or "AFTER MIDNIGHT" in str(am_label).upper():
        am_val = _cell(ws, 35, 3, L)
        if am_val:
            text_sections["after_midnight"] = _clean(am_val)

    # Situation
    sit_label = _cell(ws, 40, 1, L) or ""
    if "SITUATION" in str(sit_label).upper():
        sit_val = _cell(ws, 40, 4, L) or _cell(ws, 40, 5, L) or _cell(ws, 40, 3, L)
        if sit_val:
            v = _clean(sit_val)
            text_sections["current_operation"] = v
            text_sections["day_summary"]       = v

    # Programme prévu — value usually in col D or B
    # The label cell is typically merged across the row, so we have to verify
    # that what we read from the value column is actually different text from
    # the label itself.
    plan_label = _cell(ws, 41, 1, L) or ""
    if "PROGRAMME" in str(plan_label).upper():
        plan_val = _cell(ws, 41, 4, L) or _cell(ws, 41, 5, L) or _cell(ws, 41, 2, L)
        if plan_val and _clean(plan_val) != _clean(plan_label):
            text_sections["plan_operations"] = _clean(plan_val)

    # Remarks / Observations at row 26
    obs_label = _cell(ws, 26, 15, L) or ""
    if "OBSERVATION" in str(obs_label).upper() or "REMARK" in str(obs_label).upper():
        obs_val = _cell(ws, 26, 16, L) or _cell(ws, 27, 15, L)
        if obs_val and _clean(obs_val) != _clean(obs_label):
            text_sections["remarks"] = _clean(obs_val)

    # =====================================================================
    # SAFETY (no accident-free counter in this layout)
    # =====================================================================
    safety = {}

    wb.close()

    return {
        "header": header,
        "activities": activities,
        "text_sections": text_sections,
        "mud_checks": mud_checks,
        "mud_volume": mud_volume,
        "mud_chemical_usage": chemicals,
        "personnel_data": personnel,
        "pumps": [],
        "well_location": {},
        "survey_data": [],
        "safety": safety,
        "tarif_totals": {},
    }


# Drop-in compat
parse_daily_excel_report = parse_gw29


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        sys.exit("Usage: gw29_extract.py SOURCE.xlsx")
    data = parse_gw29(Path(sys.argv[1]))

    def default(o):
        if isinstance(o, (date_type, datetime)): return o.isoformat()
        if isinstance(o, time): return o.strftime("%H:%M:%S")
        if isinstance(o, timedelta): return o.total_seconds()
        return str(o)
    print(json.dumps(data, indent=2, default=default, ensure_ascii=False))
